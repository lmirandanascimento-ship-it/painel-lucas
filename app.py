import streamlit as st
from supabase import create_client
import pandas as pd
import plotly.graph_objects as go
import requests
import yfinance as yf
import re
import os
import unicodedata
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

# ─── Configuração ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="3P Finanças | Painel Lucas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

VERDE        = "#1A4731"
VERDE2       = "#2D6A4F"
OURO         = "#B8860B"
CAPITAL_BASE = 684_160.69
BRAPI_TOKEN  = "o1ikT8zCSyqQUkNYz224ho"
INTL_CLASSES = {"ETF USA", "REITs", "Stocks"}
CLASS_COLORS = {
    "Ações BR": "#22C55E", "ETF BR": "#16A34A", "FII": "#3B82F6",
    "ETF USA":  "#F59E0B", "REITs":  "#F97316", "Stocks": "#EF4444",
    "CDB": "#8B5CF6", "LCI/LCA": "#06B6D4", "CRI/CRA": "#0EA5E9",
    "Fundos": "#EC4899", "Tesouro Direto": "#D97706",
}

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
/* ── Base ── */
.stApp {{ background:#FFFBEF; }}
#MainMenu {{visibility:hidden;}} footer {{visibility:hidden;}}

/* ── Header ── */
.header-3p {{
    background:{VERDE}; padding:14px 28px; border-radius:10px;
    margin-bottom:18px; display:flex; justify-content:space-between; align-items:center;
}}

/* ── KPI Cards ── */
.kpi-card {{
    background:white; border:1px solid #E8E4D0;
    border-left:5px solid {VERDE}; border-radius:8px; padding:16px 18px;
}}
.kpi-card.ouro  {{ border-left-color:{OURO}; }}
.kpi-card.neg   {{ border-left-color:#B71C1C; }}
.kpi-value  {{ font-size:1.6rem; font-weight:800; color:#1A1A1A; }}
.kpi-label  {{ font-size:0.78rem; color:#777; margin-top:5px; }}
.kpi-sub    {{ font-size:0.75rem; color:#999; margin-top:3px; }}
.alerta-ouro {{
    background:#FFFBEB; border-left:4px solid {OURO};
    padding:12px 16px; border-radius:6px; margin-top:10px;
}}
.pos {{ color:#1B7A34; font-weight:700; }}
.neg {{ color:#B71C1C; font-weight:700; }}

/* ── Botões EDITAR/EXCLUIR usam <a> HTML com inline styles — sem CSS necessário ── */
</style>
""", unsafe_allow_html=True)


# ─── Supabase ─────────────────────────────────────────────────────────────────
def _secret(key: str):
    """Lê de st.secrets (Streamlit Cloud) com fallback para variável de ambiente
    (Railway, VPS ou qualquer host que não use .streamlit/secrets.toml)."""
    try:
        return st.secrets[key]
    except Exception:
        return os.environ.get(key)

@st.cache_resource
def get_sb():
    return create_client(_secret("SUPABASE_URL"), _secret("SUPABASE_KEY"))

sb = get_sb()


# ─── Utilitários ──────────────────────────────────────────────────────────────
def agora_br() -> datetime:
    """Hora atual no fuso de Brasília — o servidor (Railway) roda em UTC."""
    return datetime.now(ZoneInfo("America/Sao_Paulo"))

def brl(v, sign=False):
    if v is None: return "—"
    try:
        v = float(v)
    except Exception:
        return "—"
    if v != v:  # NaN
        return "—"
    s = f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    prefix = ("+" if v >= 0 else "-") if sign else ("-" if v < 0 else "")
    return f"R$ {prefix}{s}"

def usd(v, sign=False):
    if v is None: return "—"
    try:
        v = float(v)
    except Exception:
        return "—"
    if v != v:  # NaN
        return "—"
    prefix = ("+" if v >= 0 else "-") if sign else ("-" if v < 0 else "")
    return f"US$ {prefix}{abs(v):,.2f}"

def brl_input(v: float) -> str:
    """Formata float para string de input no padrão BRL: 40.000,00"""
    s = f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    return s

def parse_brl(s: str) -> float:
    """Converte string BRL (ex: '40.000,00' ou '800,00') para float."""
    try:
        return float(str(s).strip()
                     .replace("R$", "").replace(" ", "")
                     .replace(".", "").replace(",", "."))
    except Exception:
        return 0.0

def brl_md(v, sign=False) -> str:
    """brl() com $ escapado para uso seguro em st.markdown (evita LaTeX)."""
    return brl(v, sign=sign).replace("$", "&#36;")

# ─── Importação de planilhas de Empréstimos ("padrão Títulos") ────────────────
def _cell_str(v) -> str:
    return str(v).strip() if v is not None else ""

def _to_float_imp(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    try:
        return round(float(str(v).replace(".", "").replace(",", ".")), 2)
    except Exception:
        return 0.0

def _to_date_imp(v):
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month"):  # date
        return v
    return None

def _to_dia_imp(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        iv = int(v)
        return iv if 1 <= iv <= 31 else None
    return None

def parse_planilha_titulos(file_bytes: bytes) -> dict:
    """
    Detecta e extrai tabelas no formato "Títulos" (usado nas planilhas de
    Empréstimos Concedidos: colunas Títulos | Data Emp. | Valor Originário |
    Saldo Devedor | Parcela Juros | Recorrência | Data Prevista, com uma
    segunda tabela de pagamentos ao lado: Títulos | Data Pag. | Valor Pagamento).
    Retorna {nome_aba: {"devedor_sugerido", "emprestimos", "pagamentos_raw"}}
    apenas para as abas em que o formato foi reconhecido.
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    resultado = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, title_col, pag_col = None, None, None
        _scan_limit = min(ws.max_row, 15)
        for r in range(1, _scan_limit + 1):
            for c in range(1, ws.max_column + 1):
                _v0 = _cell_str(ws.cell(row=r, column=c).value)
                if _v0 != "Títulos":
                    continue
                _v1 = _cell_str(ws.cell(row=r, column=c + 1).value)
                _v2 = _cell_str(ws.cell(row=r, column=c + 2).value)
                if _v1 == "Data Emp." and _v2 == "Valor Originário" and title_col is None:
                    header_row, title_col = r, c
                elif _v1 == "Data Pag." and _v2 == "Valor Pagamento":
                    pag_col = c
        if header_row is None:
            continue

        # ── Nome do devedor sugerido: linha "Tabela Recebíveis/Empréstimos <Nome>" ──
        devedor_sugerido = ""
        for r in range(max(1, header_row - 3), header_row):
            for c in range(1, ws.max_column + 1):
                v = _cell_str(ws.cell(row=r, column=c).value)
                m = re.search(r"Empréstimos\s+(.+)$", v)
                if m:
                    devedor_sugerido = m.group(1).strip()

        # ── Empréstimos ──
        emprestimos = []
        for r in range(header_row + 1, ws.max_row + 1):
            titulo = _cell_str(ws.cell(row=r, column=title_col).value)
            if not titulo:
                continue
            emprestimos.append({
                "titulo":          titulo,
                "data_emprestimo": _to_date_imp(ws.cell(row=r, column=title_col + 1).value),
                "valor_original":  _to_float_imp(ws.cell(row=r, column=title_col + 2).value),
                "saldo_devedor":   _to_float_imp(ws.cell(row=r, column=title_col + 3).value),
                "parcela_juros":   _to_float_imp(ws.cell(row=r, column=title_col + 4).value),
                "recorrencia":     _cell_str(ws.cell(row=r, column=title_col + 5).value) or "Mensal",
                "dia_vencimento":  _to_dia_imp(ws.cell(row=r, column=title_col + 6).value),
            })

        # ── Pagamentos (tabela ao lado, se existir) ──
        pagamentos_raw = []
        if pag_col:
            for r in range(header_row + 1, ws.max_row + 1):
                titulo_pag = _cell_str(ws.cell(row=r, column=pag_col).value)
                if not titulo_pag:
                    continue
                pagamentos_raw.append({
                    "titulo_pag":     titulo_pag,
                    "data_pagamento": _to_date_imp(ws.cell(row=r, column=pag_col + 1).value),
                    "valor_pago":     _to_float_imp(ws.cell(row=r, column=pag_col + 2).value),
                })

        if emprestimos:
            resultado[sheet_name] = {
                "devedor_sugerido": devedor_sugerido,
                "emprestimos":      emprestimos,
                "pagamentos_raw":   pagamentos_raw,
            }
    return resultado

def match_pagamentos_titulos(titulos: list, pagamentos_raw: list) -> tuple:
    """Concilia pagamentos_raw (titulo_pag) com a lista de títulos de empréstimo,
    por igualdade exata e, como fallback, por substring (após remover 'Emp. ')."""
    matched, sem_match = [], []
    for p in pagamentos_raw:
        tp = p["titulo_pag"]
        alvo = None
        if tp in titulos:
            alvo = tp
        else:
            tp_strip = tp.replace("Emp. ", "").replace("Emp.", "").strip()
            for t in titulos:
                t_strip = t.replace("Emp. ", "").replace("Emp.", "").strip()
                if tp_strip == t_strip or tp_strip in t_strip or t_strip in tp_strip:
                    alvo = t
                    break
        if alvo:
            matched.append({**p, "titulo_emprestimo": alvo})
        else:
            sem_match.append(p)
    return matched, sem_match

def pct(v, sign=True):
    if v is None: return "—"
    v = float(v)
    s = f"{abs(v):.2f}%".replace(".",",")
    prefix = ("+" if v >= 0 else "-") if sign else ""
    return f"{prefix}{s}"

def cor_pct(v):
    if v is None: return ""
    return "pos" if float(v) >= 0 else "neg"

def html_pct(v):
    if v is None: return "—"
    cls = cor_pct(v)
    return f'<span class="{cls}">{pct(v)}</span>'


# ─── Autenticação ─────────────────────────────────────────────────────────────
def pagina_login():
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        _dbg_url = _secret("SUPABASE_URL") or ""
        _dbg_key = _secret("SUPABASE_KEY") or ""
        _dbg_key_masked = (f"{_dbg_key[:6]}...{_dbg_key[-6:]} (len={len(_dbg_key)})"
                            if _dbg_key else "(vazia/None)")
        st.caption(f"🔧 debug: SUPABASE_URL = {_dbg_url!r}")
        st.caption(f"🔧 debug: SUPABASE_KEY = {_dbg_key_masked}")
        st.markdown(f"""
        <div style='text-align:center; padding:40px 0 20px'>
            <h1 style='color:{VERDE}; font-size:2rem; margin-bottom:4px'>3P FINANÇAS</h1>
            <p style='color:{OURO}; font-style:italic; font-size:.95rem; margin:0'>
                Planejar · Poupar · Prosperar
            </p>
        </div>""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        email = st.text_input("E-mail", placeholder="seu@email.com")
        senha = st.text_input("Senha", type="password")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Entrar", use_container_width=True, type="primary"):
            if not email or not senha:
                st.warning("Preencha e-mail e senha.")
                return
            try:
                r = sb.auth.sign_in_with_password({"email": email, "password": senha})
                st.session_state["user"] = r.user
                st.rerun()
            except Exception as e:
                st.error(f"Não foi possível entrar — motivo real: {e}")


# ─── Carga de dados ───────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def load_snapshot(tipo: str) -> dict:
    """Retorna o snapshot mais recente (RV ou RF) como dict."""
    r = (sb.table("carteira_snapshots")
           .select("data,dados,usd_brl,total_investido,total_atual")
           .eq("tipo", tipo)
           .order("data", desc=True)
           .limit(1)
           .execute())
    if r.data:
        return r.data[0]
    return {}

@st.cache_data(ttl=300, show_spinner=False)
def load_posicoes_rv() -> pd.DataFrame:
    """Posições RV mais recentes (para cotações ao vivo)."""
    r = (sb.table("carteira_rv_posicoes")
           .select("*")
           .order("data_snapshot", desc=True)
           .limit(200)
           .execute())
    if not r.data:
        return pd.DataFrame()
    df = pd.DataFrame(r.data)
    # Apenas a data mais recente
    if not df.empty:
        df = df[df["data_snapshot"] == df["data_snapshot"].max()]
    return df

@st.cache_data(ttl=300, show_spinner=False)
def load_historico() -> pd.DataFrame:
    r = (sb.table("carteira_historico")
           .select("*")
           .order("data")
           .execute())
    if not r.data:
        return pd.DataFrame()
    df = pd.DataFrame(r.data)
    df["data"] = pd.to_datetime(df["data"])
    return df

@st.cache_data(ttl=300, show_spinner=False)
def load_emprestimos() -> pd.DataFrame:
    r = sb.table("emprestimos").select("*").eq("status","ativo").order("credor").execute()
    return pd.DataFrame(r.data) if r.data else pd.DataFrame()

@st.cache_data(ttl=300, show_spinner=False)
def load_investimentos() -> pd.DataFrame:
    r = sb.table("investimentos_escritorio").select("*").order("mes").execute()
    df = pd.DataFrame(r.data) if r.data else pd.DataFrame()
    if not df.empty:
        df["mes"] = pd.to_datetime(df["mes"])
    return df

@st.cache_data(ttl=300, show_spinner=False)
def load_devedores() -> pd.DataFrame:
    r = sb.table("devedores").select("*").eq("ativo", True).order("nome").execute()
    return pd.DataFrame(r.data) if r.data else pd.DataFrame()

@st.cache_data(ttl=300, show_spinner=False)
def load_emprestimos_concedidos() -> pd.DataFrame:
    r = (sb.table("emprestimos_concedidos")
           .select("*, devedores(nome, categoria)")
           .order("devedor_id")
           .order("status")
           .order("data_emprestimo")
           .execute())
    if not r.data:
        return pd.DataFrame()
    rows = []
    for item in r.data:
        dev = item.pop("devedores", {}) or {}
        item["devedor_nome"]     = dev.get("nome", "")
        item["devedor_categoria"]= dev.get("categoria", "")
        rows.append(item)
    df = pd.DataFrame(rows)
    if "data_emprestimo" in df.columns:
        df["data_emprestimo"] = pd.to_datetime(df["data_emprestimo"], errors="coerce")
    return df

@st.cache_data(ttl=300, show_spinner=False)
def load_pagamentos_recebidos(emprestimo_id: int = None) -> pd.DataFrame:
    q = sb.table("pagamentos_recebidos").select("*").order("data_pagamento", desc=True)
    if emprestimo_id:
        q = q.eq("emprestimo_id", emprestimo_id)
    r = q.execute()
    if not r.data:
        return pd.DataFrame()
    df = pd.DataFrame(r.data)
    df["data_pagamento"] = pd.to_datetime(df["data_pagamento"])
    return df

@st.cache_data(ttl=600, show_spinner=False)
def fetch_usd_brl() -> float:
    """Retorna cotação USD/BRL."""
    try:
        r = requests.get("https://economia.awesomeapi.com.br/json/last/USD-BRL", timeout=5)
        return float(r.json()["USDBRL"]["bid"])
    except Exception:
        pass
    try:
        t = yf.Ticker("USDBRL=X")
        h = t.history(period="1d")
        if not h.empty:
            return float(h["Close"].iloc[-1])
    except Exception:
        pass
    return 5.75

@st.cache_data(ttl=600, show_spinner=False)
def fetch_precos_br(tickers_br: tuple) -> dict:
    """Retorna preços de ativos BR via BRAPI, com fallback para yfinance (.SA)."""
    prices = {}
    if not tickers_br:
        return prices
    # Tentativa 1: BRAPI
    try:
        url = f"https://brapi.dev/api/quote/{','.join(tickers_br)}?token={BRAPI_TOKEN}"
        resp = requests.get(url, timeout=15)
        if resp.ok:
            for item in resp.json().get("results", []):
                if item and item.get("regularMarketPrice"):
                    sym = item["symbol"].replace(".SA", "")
                    prices[sym] = item["regularMarketPrice"]
                    prices[item["symbol"]] = item["regularMarketPrice"]
    except Exception:
        pass
    # Fallback: yfinance para os que não vieram do BRAPI
    missing = [t for t in tickers_br
               if t not in prices and t.replace(".SA", "") not in prices]
    if missing:
        try:
            tks_sa  = [t + ".SA" if not t.endswith(".SA") else t for t in missing]
            tks_str = " ".join(tks_sa)
            data_yf = yf.download(tks_str, period="2d", auto_adjust=True, progress=False)
            if not data_yf.empty:
                close = (data_yf["Close"] if "Close" in data_yf.columns
                         else data_yf.xs("Close", axis=1, level=0))
                if hasattr(close, "columns"):
                    for tk, tk_sa in zip(missing, tks_sa):
                        if tk_sa in close.columns:
                            v = close[tk_sa].dropna()
                            if not v.empty:
                                prices[tk] = float(v.iloc[-1])
                else:
                    v = close.dropna()
                    if not v.empty and len(missing) == 1:
                        prices[missing[0]] = float(v.iloc[-1])
        except Exception:
            pass
    return prices

@st.cache_data(ttl=600, show_spinner=False)
def fetch_precos_us(tickers_us: tuple) -> dict:
    """Retorna preços de ativos US via yfinance."""
    prices = {}
    if not tickers_us:
        return prices
    try:
        tks_str = " ".join(tickers_us)
        data = yf.download(tks_str, period="2d", auto_adjust=True, progress=False)
        if not data.empty:
            close = data["Close"] if "Close" in data.columns else data.xs("Close", axis=1, level=0)
            if hasattr(close, "columns"):
                for tk in tickers_us:
                    if tk in close.columns:
                        v = close[tk].dropna()
                        if not v.empty:
                            prices[tk] = float(v.iloc[-1])
            else:
                v = close.dropna()
                if not v.empty and len(tickers_us) == 1:
                    prices[tickers_us[0]] = float(v.iloc[-1])
    except Exception:
        pass
    return prices

# mantido para compatibilidade de assinatura em locais que chamam com dois args
@st.cache_data(ttl=600, show_spinner=False)
def fetch_precos_brapi(tickers_br: tuple, tickers_us: tuple) -> tuple:
    """Retorna (dict_precos_br+us, usd_brl)."""
    prices = fetch_precos_br(tickers_br)
    prices.update(fetch_precos_us(tickers_us))
    usd_brl = fetch_usd_brl()
    return prices, usd_brl


# ─── Header ───────────────────────────────────────────────────────────────────
def render_header():
    user = st.session_state.get("user")
    nome = (user.email or "").split("@")[0].capitalize()
    st.markdown(f"""
    <div class='header-3p'>
        <div>
            <span style='color:{OURO};font-size:1.2rem;font-weight:800'>3P FINANÇAS</span>
            <span style='color:#bbb;font-size:.85rem;margin-left:14px'>Planejar · Poupar · Prosperar</span>
        </div>
        <div style='color:#ddd;font-size:.82rem'>
            📅 {agora_br().strftime("%d/%m/%Y")} &nbsp;|&nbsp; 👤 {nome}
        </div>
    </div>""", unsafe_allow_html=True)


# ─── TAB: RESUMO ──────────────────────────────────────────────────────────────
def tab_resumo(snap_rv, snap_rf, posicoes_rv):
    dados_rv = snap_rv.get("dados", {}) if snap_rv else {}
    dados_rf = snap_rf.get("dados", {}) if snap_rf else {}
    rv_tot   = float(snap_rv.get("total_atual", 0) or 0)
    rf_tot   = float(snap_rf.get("total_atual", 0) or 0)
    total    = rv_tot + rf_tot
    ganho    = total - CAPITAL_BASE
    rentab   = ganho / CAPITAL_BASE * 100 if CAPITAL_BASE else 0
    usd_brl  = float(snap_rv.get("usd_brl", 5.75) or 5.75)

    # Botão cotações ao vivo
    c_btn, c_status = st.columns([2, 6])
    with c_btn:
        if st.button("🔄 Atualizar Cotações", type="secondary"):
            with st.spinner("Buscando cotações..."):
                st.cache_data.clear()
                ok_list, fail_list = [], []

                # — USD/BRL —
                try:
                    r = requests.get("https://economia.awesomeapi.com.br/json/last/USD-BRL", timeout=5)
                    usd_val = float(r.json()["USDBRL"]["bid"])
                    ok_list.append(f"USD/BRL → {usd_val:.4f}")
                except Exception:
                    try:
                        h = yf.Ticker("USDBRL=X").history(period="1d")
                        usd_val = float(h["Close"].iloc[-1])
                        ok_list.append(f"USD/BRL → {usd_val:.4f} (yfinance)")
                    except Exception:
                        fail_list.append("USD/BRL")

                # — Tickers BR —
                tickers_br_up, tickers_us_up = [], []
                if not posicoes_rv.empty:
                    tickers_br_up = posicoes_rv[posicoes_rv["moeda"] == "BRL"]["ticker"].tolist()
                    tickers_us_up = posicoes_rv[posicoes_rv["moeda"] == "USD"]["ticker"].tolist()

                if tickers_br_up:
                    found_br = {}
                    # Tentativa 1: BRAPI
                    try:
                        url  = f"https://brapi.dev/api/quote/{','.join(tickers_br_up)}?token={BRAPI_TOKEN}"
                        resp = requests.get(url, timeout=15)
                        if resp.ok:
                            for item in resp.json().get("results", []):
                                if item and item.get("regularMarketPrice"):
                                    sym = item["symbol"].replace(".SA", "")
                                    found_br[sym] = (item["regularMarketPrice"], "BRAPI")
                    except Exception:
                        pass
                    # Fallback: yfinance para os que não vieram do BRAPI
                    missing_br = [t for t in tickers_br_up
                                  if t.replace(".SA", "") not in found_br]
                    if missing_br:
                        try:
                            tks_sa  = [t + ".SA" if not t.endswith(".SA") else t for t in missing_br]
                            data_yf = yf.download(" ".join(tks_sa), period="2d",
                                                   auto_adjust=True, progress=False)
                            if not data_yf.empty:
                                close = (data_yf["Close"] if "Close" in data_yf.columns
                                         else data_yf.xs("Close", axis=1, level=0))
                                if hasattr(close, "columns"):
                                    for tk, tk_sa in zip(missing_br, tks_sa):
                                        if tk_sa in close.columns:
                                            v = close[tk_sa].dropna()
                                            if not v.empty:
                                                found_br[tk.replace(".SA","")] = (float(v.iloc[-1]), "yfinance")
                                else:
                                    v = close.dropna()
                                    if not v.empty and len(missing_br) == 1:
                                        tk = missing_br[0].replace(".SA","")
                                        found_br[tk] = (float(v.iloc[-1]), "yfinance")
                        except Exception:
                            pass
                    for tk in tickers_br_up:
                        tk_c = tk.replace(".SA", "")
                        if tk_c in found_br:
                            preco, fonte = found_br[tk_c]
                            fonte_str = f" ({fonte})" if fonte != "BRAPI" else ""
                            ok_list.append(f"{tk_c} → R$ {preco:.2f}{fonte_str}")
                        else:
                            fail_list.append(tk_c)

                # — Tickers US —
                if tickers_us_up:
                    try:
                        tks_str = " ".join(tickers_us_up)
                        data_yf = yf.download(tks_str, period="2d", auto_adjust=True, progress=False)
                        if not data_yf.empty:
                            close = (data_yf["Close"] if "Close" in data_yf.columns
                                     else data_yf.xs("Close", axis=1, level=0))
                            if hasattr(close, "columns"):
                                for tk in tickers_us_up:
                                    if tk in close.columns:
                                        v = close[tk].dropna()
                                        if not v.empty:
                                            ok_list.append(f"{tk} → USD {float(v.iloc[-1]):.2f}")
                                        else:
                                            fail_list.append(tk)
                                    else:
                                        fail_list.append(tk)
                            else:
                                v = close.dropna()
                                if not v.empty and len(tickers_us_up) == 1:
                                    ok_list.append(f"{tickers_us_up[0]} → USD {float(v.iloc[-1]):.2f}")
                                else:
                                    fail_list.extend(tickers_us_up)
                        else:
                            fail_list.extend(tickers_us_up)
                    except Exception:
                        fail_list.extend(tickers_us_up)

                st.session_state["ultima_atualizacao"] = {
                    "ok":   ok_list,
                    "fail": fail_list,
                    "hora": agora_br().strftime("%H:%M:%S"),
                }
            st.rerun()

    with c_status:
        data_rv = snap_rv.get("data","—") if snap_rv else "—"
        usd_brl_hoje = fetch_usd_brl()
        st.caption(
            f"RV: snapshot de {data_rv} · RF: snapshot de {snap_rf.get('data','—') if snap_rf else '—'} · "
            f"USD/BRL no snapshot: {usd_brl:.4f} · USD/BRL hoje: {usd_brl_hoje:.4f}"
        )

    # ── Resultado da última atualização ─────────────────────────────────────────
    if "ultima_atualizacao" in st.session_state:
        res    = st.session_state["ultima_atualizacao"]
        n_ok   = len(res["ok"])
        n_fail = len(res["fail"])
        cor    = "🟢" if n_fail == 0 else ("🟡" if n_ok > 0 else "🔴")
        label  = f"{cor} Atualização das {res['hora']} — {n_ok} atualizados · {n_fail} com falha"
        with st.expander(label, expanded=(n_fail > 0)):
            col_ok, col_fail = st.columns(2)
            with col_ok:
                st.markdown("**✅ Atualizados com sucesso**")
                for item in res["ok"]:
                    st.markdown(f"- {item}")
            if res["fail"]:
                with col_fail:
                    st.markdown("**❌ Falhou (usando valor do snapshot)**")
                    for item in res["fail"]:
                        st.markdown(f"- {item}")

    # KPIs
    c1, c2, c3, c4, c5 = st.columns(5)
    def kpi(col, label, valor, sub="", cls=""):
        col.markdown(f"""
        <div class='kpi-card {cls}'>
            <div class='kpi-value'>{valor}</div>
            <div class='kpi-label'>{label}</div>
            {'<div class="kpi-sub">'+sub+'</div>' if sub else ''}
        </div>""", unsafe_allow_html=True)

    kpi(c1, "Patrimônio Total",        brl(total), "Valor de mercado atual")
    kpi(c2, "Capital Inicial (ref.)",  brl(CAPITAL_BASE), "Mar/2026", "ouro")
    cls_g = "" if ganho >= 0 else "neg"
    kpi(c3, "Lucro Acumulado",         brl(ganho, sign=True), pct(rentab), cls_g)
    kpi(c4, "Renda Variável (snapshot)",brl(rv_tot), f"Ref. {data_rv}")
    kpi(c5, "Renda Fixa (snapshot)",   brl(rf_tot))

    st.markdown("<br>", unsafe_allow_html=True)

    # Pizza composição
    st.markdown(f"#### 🍕 Composição do Patrimônio")
    classes_vals = {}
    for cls, data in dados_rv.get("classes", {}).items():
        v = data.get("atual", 0) or 0
        if v > 0: classes_vals[cls] = v
    for cls, data in dados_rf.get("classes", {}).items():
        v = data.get("atual", 0) or 0
        if v > 0: classes_vals[cls] = round(classes_vals.get(cls, 0) + v, 2)

    if classes_vals:
        labels = list(classes_vals.keys())
        values = [classes_vals[k] for k in labels]
        colors = [CLASS_COLORS.get(k, "#999") for k in labels]
        fig = go.Figure(go.Pie(
            labels=labels, values=values, hole=0.6,
            marker_colors=colors,
            textinfo="percent",
            textfont_size=12,
        ))
        fig.update_layout(
            height=340, margin=dict(l=0,r=0,t=0,b=0),
            paper_bgcolor="#FFFBEF",
            legend=dict(orientation="v", x=1.0, y=0.5),
            annotations=[dict(text=brl(total), x=0.5, y=0.5,
                              font_size=14, showarrow=False)],
        )
        c_pizza, c_leg = st.columns([1.3, 1])
        with c_pizza:
            st.plotly_chart(fig, use_container_width=True)
        with c_leg:
            st.markdown("<br>", unsafe_allow_html=True)
            sorted_cls = sorted(classes_vals.items(), key=lambda x: -x[1])
            for k, v in sorted_cls:
                pct_v = v / total * 100 if total else 0
                cor = CLASS_COLORS.get(k, "#999")
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:10px;padding:4px 0">'
                    f'<span style="width:12px;height:12px;background:{cor};border-radius:3px;flex-shrink:0;display:inline-block"></span>'
                    f'<span style="flex:1;font-size:13px">{k}</span>'
                    f'<span style="font-size:12px;color:#777">{pct_v:.1f}%</span>'
                    f'<span style="font-size:13px;font-weight:700;min-width:110px;text-align:right">{brl(v)}</span>'
                    f'</div>', unsafe_allow_html=True)


# ─── TAB: Ativos RV BR (Ações, ETF, FII) ─────────────────────────────────────
def tab_rv_br(classe: str, snap_rv, posicoes_rv):
    dados = snap_rv.get("dados", {}) if snap_rv else {}
    cls_data = dados.get("classes", {}).get(classe, {})
    posicoes = cls_data.get("posicoes", [])

    if not posicoes:
        st.info("Nenhuma posição nesta classe no snapshot atual.")
        return

    # Buscar cotações ao vivo
    tickers = [p.get("nome","") for p in posicoes if p.get("nome")]
    prices, usd_brl = {}, 5.75
    if not posicoes_rv.empty:
        pos_cls = posicoes_rv[posicoes_rv["classe"] == classe]
        tickers_br = tuple(pos_cls["ticker"].tolist())
        if tickers_br:
            with st.spinner("Buscando cotações..."):
                prices, usd_brl = fetch_precos_brapi(tickers_br, ())

    # Montar tabela
    rows = []
    tot_inv = 0.0; tot_at = 0.0
    for p in posicoes:
        nome    = p.get("nome", "")
        inv     = float(p.get("investido") or 0)
        at_snap = float(p.get("atual") or 0)

        # Preço ao vivo
        ticker_live = nome.replace(".SA","") if nome.endswith(".SA") else nome
        preco_live  = prices.get(ticker_live) or prices.get(nome)
        qtd_val     = float(p.get("qtd") or 0)
        at_live     = preco_live * qtd_val if preco_live and qtd_val else at_snap
        ren = (at_live / inv - 1) * 100 if inv else 0

        tot_inv += inv
        tot_at  += at_live

        rows.append({
            "Ativo":         nome,
            "Setor":         p.get("setor",""),
            "Qtd":           f"{qtd_val:g}" if qtd_val else "—",
            "PM (R$)":       brl(p.get("preco_pago_brl") or (inv / qtd_val if qtd_val else None)),
            "Cotação":       brl(preco_live) if preco_live else "⟳",
            "Investido":     brl(inv),
            "Posição Atual": brl(at_live),
            "Ganho":         brl(at_live - inv, sign=True),
            "%":             pct(ren),
            "_ren":          ren,
        })

    # Total
    ren_tot = (tot_at / tot_inv - 1) * 100 if tot_inv else 0
    rows.append({
        "Ativo": "TOTAL", "Setor": "", "Qtd": "", "PM (R$)": "",
        "Cotação": "", "Investido": brl(tot_inv), "Posição Atual": brl(tot_at),
        "Ganho": brl(tot_at - tot_inv, sign=True), "%": pct(ren_tot), "_ren": ren_tot,
    })

    df = pd.DataFrame(rows)
    df_show = df.drop(columns=["_ren"])

    # Destacar linhas positivas/negativas
    def highlight(row):
        if row["Ativo"] == "TOTAL":
            return ["background:#F0EDD8; font-weight:700"] * len(row)
        ren = df.loc[row.name, "_ren"]
        if ren <= -10: bg = "#FFF9C4"
        elif ren >= 10: bg = "#C8E6C9"
        else: bg = ""
        return [f"background:{bg}" if bg else ""] * len(row)

    st.dataframe(df_show.style.apply(highlight, axis=1),
                 use_container_width=True, hide_index=True, height=350)

    st.caption(f"Cotações: BRAPI ao vivo | Snapshot: {snap_rv.get('data','—')}")


# ─── TAB: Ativos Internacionais ───────────────────────────────────────────────
def tab_internacional(snap_rv, posicoes_rv):
    dados = snap_rv.get("dados", {}) if snap_rv else {}
    usd_brl = float(snap_rv.get("usd_brl", 5.75) or 5.75)

    # Buscar tickers US
    pos_us = pd.DataFrame()
    if not posicoes_rv.empty:
        pos_us = posicoes_rv[posicoes_rv["moeda"] == "USD"]
    tickers_us = tuple(pos_us["ticker"].tolist()) if not pos_us.empty else ()
    prices = {}
    if tickers_us:
        with st.spinner("Buscando cotações internacionais..."):
            prices, usd_brl = fetch_precos_brapi((), tickers_us)

    tabs_intl = st.tabs(["📊 ETF USA", "🏠 REITs", "📈 Stocks"])

    for tab_obj, classe in zip(tabs_intl, ["ETF USA", "REITs", "Stocks"]):
        with tab_obj:
            cls_data = dados.get("classes", {}).get(classe, {})
            posicoes = cls_data.get("posicoes", [])
            if not posicoes:
                st.info("Sem posições.")
                continue

            rows = []
            tot_inv_usd = 0.0; tot_at_usd = 0.0
            for p in posicoes:
                nome        = p.get("nome", "")
                qtd         = float(p.get("qtd") or 0)
                pm_usd      = float(p.get("preco_pago_usd") or p.get("preco_atual_usd") or 0)
                p_live      = prices.get(nome)
                cotacao_usd = p_live if p_live else float(p.get("preco_atual_usd") or 0)
                inv_usd     = round(qtd * pm_usd, 2)
                at_usd      = round(qtd * cotacao_usd, 2)
                ren         = (at_usd / inv_usd - 1) * 100 if inv_usd else 0
                tot_inv_usd += inv_usd; tot_at_usd += at_usd

                rows.append({
                    "Ativo":            nome,
                    "Setor":            p.get("setor",""),
                    "Qtd":              f"{qtd:g}" if qtd else "—",
                    "PM (US$)":         usd(pm_usd) if pm_usd else "—",
                    "Cotação (US$)":    usd(cotacao_usd) if p_live else "⟳",
                    "Investido (US$)":  usd(inv_usd),
                    "Posição (US$)":    usd(at_usd),
                    "Ganho (US$)":      usd(at_usd - inv_usd, sign=True),
                    "%":                pct(ren),
                    "Posição (R$) est.": brl(at_usd * usd_brl),
                })
            ren_t = (tot_at_usd / tot_inv_usd - 1) * 100 if tot_inv_usd else 0
            rows.append({"Ativo":"TOTAL","Setor":"","Qtd":"","PM (US$)":"",
                         "Cotação (US$)":"","Investido (US$)":usd(tot_inv_usd),
                         "Posição (US$)":usd(tot_at_usd),
                         "Ganho (US$)":usd(tot_at_usd-tot_inv_usd,sign=True),
                         "%":pct(ren_t), "Posição (R$) est.": brl(tot_at_usd * usd_brl)})
            st.dataframe(pd.DataFrame(rows), use_container_width=True,
                         hide_index=True, height=300)
    st.caption(
        f"PM, Cotação, Investido, Posição, Ganho e % são 100% em dólar — a rentabilidade "
        f"não sofre efeito de câmbio. A coluna \"Posição (R$) est.\" é apenas uma "
        f"conversão da posição em dólar pelo câmbio USD/BRL de hoje ({usd_brl:.4f}) "
        f"— é uma estimativa de valor em reais no dia, não faz parte do cálculo de "
        f"rentabilidade. Cotações: BRAPI."
    )


# ─── TAB genérica RF ─────────────────────────────────────────────────────────
def _tbl_rf(posicoes, colunas, total_field="atual"):
    if not posicoes:
        st.info("Sem dados disponíveis.")
        return
    rows = []
    tot = 0.0
    for p in posicoes:
        row = {}
        for col_label, col_key in colunas:
            v = p.get(col_key)
            if col_key in ("atual","investido","ganho","valor_liquido","posicao_mercado"):
                row[col_label] = brl(v)
            elif col_key in ("rentab","rentab_liquida","rentab_bruta"):
                row[col_label] = pct((v or 0) * 100) if v is not None else "—"
            else:
                row[col_label] = str(v) if v is not None else "—"
        rows.append(row)
        tot += float(p.get(total_field) or 0)

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.metric("Total", brl(tot))


def tab_tesouro(snap_rf):
    dados = snap_rf.get("dados", {}) if snap_rf else {}
    posicoes = dados.get("classes", {}).get("Tesouro Direto", {}).get("posicoes", [])
    colunas = [
        ("Título",    "nome"),
        ("Qtd",       "qtd"),
        ("Vencimento","vencimento"),
        ("Taxa",      "tipo_taxa"),
        ("Investido", "investido"),
        ("Rentab.",   "rentab_liquida"),
        ("Valor Líq.","valor_liquido"),
    ]
    _tbl_rf(posicoes, colunas, total_field="valor_liquido")
    st.caption("IR regressivo já descontado no Valor Líquido.")

def tab_cricra(snap_rf):
    dados = snap_rf.get("dados", {}) if snap_rf else {}
    posicoes = dados.get("classes", {}).get("CRI/CRA", {}).get("posicoes", [])
    colunas = [
        ("Título",    "nome"),
        ("Taxa Mercado","taxa_mercado"),
        ("Investido", "investido"),
        ("Valor Atual","atual"),
        ("Rentab.",   "rentab"),
    ]
    _tbl_rf(posicoes, colunas)
    st.caption("CRI/CRA: isentos de IR para pessoa física.")

def tab_fundos(snap_rf):
    dados = snap_rf.get("dados", {}) if snap_rf else {}
    posicoes = dados.get("classes", {}).get("Fundos", {}).get("posicoes", [])
    colunas = [
        ("Fundo",     "nome"),
        ("Investido", "investido"),
        ("Valor Liq.","valor_liquido"),
        ("Rent. Líq.","rentab_liquida"),
    ]
    _tbl_rf(posicoes, colunas, total_field="valor_liquido")

def tab_cdb(snap_rf):
    dados = snap_rf.get("dados", {}) if snap_rf else {}
    cdb  = dados.get("classes", {}).get("CDB", {}).get("posicoes", [])
    lcis = dados.get("classes", {}).get("LCI/LCA", {}).get("posicoes", [])
    posicoes = cdb + lcis
    colunas = [
        ("Título",    "nome"),
        ("Investido", "investido"),
        ("Atual",     "atual"),
        ("Ganho",     "ganho"),
        ("Rentab.",   "rentab"),
    ]
    _tbl_rf(posicoes, colunas)
    st.caption("LCI/LCA: isentos de IR. CDB: valor bruto (IR na resgate).")


# ─── TAB: EVOLUÇÃO ────────────────────────────────────────────────────────────
def tab_evolucao(historico: pd.DataFrame):
    if historico.empty:
        st.info("Histórico disponível após a primeira atualização via executar_atualizacao.command.")
        return

    ultimo = historico.iloc[-1]
    c1, c2, c3 = st.columns(3)
    c1.metric("Último Snapshot", brl(float(ultimo["total_atual"])))
    c2.metric("Ganho Acumulado", brl(float(ultimo["total_ganho"]), sign=True))
    c3.metric("Rentabilidade", pct(float(ultimo["total_rentab"])))

    fig = go.Figure()

    # Linha de patrimônio total
    fig.add_trace(go.Scatter(
        x=historico["data"], y=historico["total_atual"],
        mode="lines+markers", name="Patrimônio Total",
        line=dict(color="#2E7D52", width=2),
        marker=dict(size=5),
        hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Patrimônio: R$ %{y:,.0f}<extra></extra>",
    ))

    # Linha de capital base
    n = len(historico)
    fig.add_trace(go.Scatter(
        x=historico["data"],
        y=[CAPITAL_BASE] * n,
        mode="lines", name="Capital Base",
        line=dict(color=OURO, width=1.5, dash="dash"),
        hoverinfo="skip",
    ))

    fig.update_layout(
        height=360,
        margin=dict(l=0, r=0, t=20, b=0),
        paper_bgcolor="#FFFBEF",
        plot_bgcolor="#FFFBEF",
        yaxis=dict(tickformat=",.0f", tickprefix="R$ "),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)

    # Tabela de histórico
    df_hist_show = historico[["data","total_atual","total_ganho","total_rentab"]].copy()
    df_hist_show = df_hist_show.sort_values("data", ascending=False)
    df_hist_show.columns = ["Data","Patrimônio","Ganho Acum.","Rentab."]
    df_hist_show["Data"]        = df_hist_show["Data"].dt.strftime("%d/%m/%Y")
    df_hist_show["Patrimônio"]  = df_hist_show["Patrimônio"].apply(brl)
    df_hist_show["Ganho Acum."] = df_hist_show["Ganho Acum."].apply(lambda v: brl(float(v), sign=True))
    df_hist_show["Rentab."]     = df_hist_show["Rentab."].apply(lambda v: pct(float(v)))
    st.dataframe(df_hist_show, use_container_width=True, hide_index=True)



# ─── helper: conteúdo por devedor ──────────────────────────────────────────────
def _conteudo_devedor(dev_id: int, dev_nome: str):
    """KPIs, tabelas, formulário e histórico de um devedor."""
    # Proteção: falha de rede/Supabase após .clear() não propaga "Oh no." para o Streamlit
    try:
        emp_conc_df = load_emprestimos_concedidos()
        pagtos_all  = load_pagamentos_recebidos()
    except Exception as _load_err:
        st.error(f"⚠️ Erro ao carregar dados. Recarregue a página. ({_load_err})")
        if st.button("🔄 Tentar novamente", key=f"retry_load_{dev_id}"):
            load_emprestimos_concedidos.clear()
            load_pagamentos_recebidos.clear()
            st.rerun()
        return

    emp_dev  = emp_conc_df[emp_conc_df["devedor_id"] == dev_id] if not emp_conc_df.empty else pd.DataFrame()
    ativos_d = emp_dev[emp_dev["status"] == "ativo"]  if not emp_dev.empty else pd.DataFrame()
    if not ativos_d.empty:
        ativos_d = ativos_d.sort_values("data_emprestimo").reset_index(drop=True)
    quit_d = emp_dev[emp_dev["status"] == "quitado"] if not emp_dev.empty else pd.DataFrame()

    pagtos_d = pagtos_all[pagtos_all["emprestimo_id"].isin(emp_dev["id"].tolist())] \
               if not pagtos_all.empty and not emp_dev.empty else pd.DataFrame()

    # ── KPIs ──────────────────────────────────────────────────────────────────
    saldo_d = float(ativos_d["saldo_devedor"].sum()) if not ativos_d.empty else 0.0
    juros_d = float(ativos_d["parcela_juros"].sum()) if not ativos_d.empty else 0.0
    receb_d = float(pagtos_d["valor_pago"].sum())    if not pagtos_d.empty else 0.0

    ka, kb, kc = st.columns(3)
    ka.markdown(f"""<div class='kpi-card'><div class='kpi-value'>{brl(saldo_d)}</div>
        <div class='kpi-label'>💰 Saldo Devedor</div></div>""", unsafe_allow_html=True)
    kb.markdown(f"""<div class='kpi-card ouro'><div class='kpi-value'>{brl(juros_d)}</div>
        <div class='kpi-label'>📅 Juros/Mês</div></div>""", unsafe_allow_html=True)
    kc.markdown(f"""<div class='kpi-card'><div class='kpi-value'>{brl(receb_d)}</div>
        <div class='kpi-label'>✅ Total Recebido</div></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Contratos ativos — cards com barra de progresso ──────────────────────
    if not ativos_d.empty:
        st.markdown(
            "<span style='font-size:0.9rem;font-weight:700;color:#1A1A1A'>"
            f"Contratos Ativos &nbsp;"
            f"<span style='font-size:0.78rem;font-weight:400;color:#888'>"
            f"({len(ativos_d)} contrato{'s' if len(ativos_d)>1 else ''})</span>"
            "</span>",
            unsafe_allow_html=True,
        )
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        _cards = []
        for _, _r in ativos_d.sort_values("dia_vencimento", na_position="last").iterrows():
            _titulo   = str(_r["titulo"])
            _orig     = float(_r["valor_original"])
            _saldo    = float(_r["saldo_devedor"])
            _juros    = float(_r["parcela_juros"])
            _dv       = _r["dia_vencimento"]
            _dia_str  = f"Dia {int(_dv)}" if pd.notna(_dv) else "—"
            _amort_v  = max(0.0, _orig - _saldo)
            _amort_p  = round(min(100.0, _amort_v / _orig * 100), 1) if _orig > 0 else 0.0
            # Cor da barra: verde quando amortizando, cinza quando sem pagamentos
            _bar_col  = VERDE2 if _amort_p > 0 else "#ccc"
            _cards.append(
                f'<div style="background:#fff;border:1px solid #E8E4D0;border-radius:12px;'
                f'padding:14px 16px;display:flex;flex-direction:column;gap:10px;">'

                # ── Título + badge de vencimento
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px;">'
                f'<span style="font-size:13px;font-weight:600;color:#1A1A1A;line-height:1.3">{_titulo}</span>'
                f'<span style="font-size:11px;color:#666;background:#F5F2E8;padding:3px 9px;'
                f'border-radius:20px;white-space:nowrap;flex-shrink:0">Vence {_dia_str}</span>'
                f'</div>'

                # ── Saldo + Juros
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">'
                f'<div style="display:flex;flex-direction:column;gap:2px;">'
                f'<span style="font-size:15px;font-weight:700;color:#1A1A1A">{brl(_saldo)}</span>'
                f'<span style="font-size:10px;color:#999;letter-spacing:.02em">Saldo devedor</span>'
                f'</div>'
                f'<div style="display:flex;flex-direction:column;gap:2px;">'
                f'<span style="font-size:15px;font-weight:700;color:{OURO}">{brl(_juros)}</span>'
                f'<span style="font-size:10px;color:#999;letter-spacing:.02em">Juros / mês</span>'
                f'</div>'
                f'</div>'

                # ── Barra de progresso
                f'<div style="display:flex;flex-direction:column;gap:4px;">'
                f'<div style="height:5px;border-radius:99px;background:#E8E4D0;overflow:hidden;">'
                f'<div style="height:100%;width:{_amort_p}%;border-radius:99px;background:{_bar_col};'
                f'transition:width .4s ease;"></div>'
                f'</div>'
                f'<div style="display:flex;justify-content:space-between;font-size:10px;color:#aaa;">'
                f'<span>{_amort_p:.0f}% amortizado &nbsp;({brl(_amort_v)})</span>'
                f'<span>Original: {brl(_orig)}</span>'
                f'</div>'
                f'</div>'

                f'</div>'
            )

        # Grid 2 colunas (1 se só tiver 1 contrato)
        _cols = "1fr 1fr" if len(_cards) > 1 else "1fr"
        st.markdown(
            f'<div style="display:grid;grid-template-columns:{_cols};gap:10px;margin-bottom:8px;">'
            + "".join(_cards)
            + "</div>",
            unsafe_allow_html=True,
        )

    # ── Quitados ───────────────────────────────────────────────────────────────
    if not quit_d.empty:
        with st.expander(f"✅ Quitados ({len(quit_d)})"):
            datas_quit = {}
            if not pagtos_d.empty:
                for _, qrow in quit_d.iterrows():
                    pq = pagtos_d[pagtos_d["emprestimo_id"] == int(qrow["id"])]
                    if not pq.empty:
                        datas_quit[int(qrow["id"])] = pq["data_pagamento"].max()

            _rows_q = []
            for _, _q in quit_d.iterrows():
                _qid      = int(_q["id"])
                _qtit     = str(_q["titulo"])
                _qorig    = float(_q["valor_original"])
                _qemp     = _q["data_emprestimo"].strftime("%d/%m/%Y") if pd.notna(_q["data_emprestimo"]) else "—"
                _qquit    = datas_quit[_qid].strftime("%d/%m/%Y") if _qid in datas_quit else "—"
                _rows_q.append(
                    f'<div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr;'
                    f'gap:8px;align-items:center;padding:8px 12px;'
                    f'border-bottom:1px solid #E8E4D0;font-size:0.85em;">'
                    f'<span style="font-weight:600;color:#1A1A1A">{_qtit}</span>'
                    f'<span style="color:#555">{brl(_qorig)}</span>'
                    f'<span style="color:#888">Emp. {_qemp}</span>'
                    f'<span style="color:#2D6A4F;font-weight:600">✓ {_qquit}</span>'
                    f'</div>'
                )
            header_q = (
                f'<div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr;'
                f'gap:8px;padding:6px 12px;background:#F5F2E8;'
                f'border-radius:8px 8px 0 0;font-size:0.75em;font-weight:600;color:#888;">'
                f'<span>Título</span><span>Valor Original</span>'
                f'<span>Data Emp.</span><span>Quitado em</span></div>'
            )
            st.markdown(
                f'<div style="border:1px solid #E8E4D0;border-radius:8px;overflow:hidden;">'
                + header_q + "".join(_rows_q)
                + "</div>",
                unsafe_allow_html=True,
            )

    st.markdown("---")

    # ── Registrar Pagamento ────────────────────────────────────────────────────
    with st.expander("💸 Registrar Pagamento Recebido"):
        if ativos_d.empty:
            st.info("Nenhum contrato ativo para este devedor.")
        else:
            # Seletor fora do form → atualiza preview sem enviar
            contratos_opts = ativos_d["titulo"].tolist()
            sel_pag = st.selectbox("Contrato", contratos_opts, key=f"pag_sel_{dev_id}")
            row_pag  = ativos_d[ativos_d["titulo"] == sel_pag].iloc[0]
            saldo_p  = float(row_pag["saldo_devedor"])
            juros_p  = float(row_pag["parcela_juros"])
            emp_id_p = int(row_pag["id"])
            taxa_p   = float(row_pag["taxa_juros"])
            st.markdown(
                f'<small style="color:#555">Saldo atual: <b>{brl_md(saldo_p)}</b> &nbsp;·&nbsp; Juros do mês: <b>{brl_md(juros_p)}</b></small>',
                unsafe_allow_html=True)

            with st.form(key=f"form_pag_{dev_id}", clear_on_submit=True):
                tipo_pag = st.radio(
                    "Tipo de pagamento",
                    ["Amortização (abate o saldo)", "Somente Juros (não abate o saldo)"],
                    horizontal=True, key=f"tipo_pag_{dev_id}",
                )
                fp1, fp2 = st.columns(2)
                valor_p_str = fp1.text_input("Valor recebido (R$)",
                                              value=brl_input(juros_p),
                                              placeholder="ex: 800,00",
                                              help="Digite o valor no formato 800,00 ou 2.000,50")
                valor_p = parse_brl(valor_p_str)
                data_p  = fp2.date_input("Data do recebimento",
                                          value=agora_br().date(),
                                          min_value=date.today() - timedelta(days=365),
                                          max_value=date.today(),
                                          format="DD/MM/YYYY")
                obs_p   = st.text_input("Observação (opcional)")
                submitted_pag = st.form_submit_button(
                    "✅ Confirmar Recebimento", type="primary", use_container_width=True)

            if submitted_pag:
                eh_somente_juros = tipo_pag.startswith("Somente")
                _hoje = date.today()
                if data_p > _hoje:
                    st.error("❌ Data futura não permitida. Utilize uma data até hoje.")
                elif data_p < _hoje - timedelta(days=365):
                    st.error("❌ Data muito antiga. O limite máximo é 1 ano no passado.")
                else:
                    if eh_somente_juros:
                        # Pagamento somente de juros: não abate o saldo devedor.
                        amort_p      = 0.0
                        novo_saldo_p = saldo_p
                        novo_juros_p = juros_p
                        juros_reg    = round(valor_p, 2)
                    else:
                        # Amortização pura: o valor pago debita integralmente do saldo.
                        # Juros do período ficam registrados como referência, sem abater do valor.
                        amort_p      = valor_p
                        novo_saldo_p = round(max(0.0, saldo_p - amort_p), 2)
                        novo_juros_p = round(novo_saldo_p * taxa_p, 2)
                        juros_reg    = round(juros_p, 2)
                    try:
                        sb.table("pagamentos_recebidos").insert({
                            "emprestimo_id":  emp_id_p,
                            "data_pagamento": str(data_p),
                            "valor_pago":     round(valor_p, 2),
                            "juros":          juros_reg,
                            "amortizacao":    round(amort_p, 2),
                            "saldo_antes":    round(saldo_p, 2),
                            "saldo_depois":   round(novo_saldo_p, 2),
                            "observacao":     obs_p,
                            "tipo":           "juros" if eh_somente_juros else "amortizacao",
                        }).execute()
                        if not eh_somente_juros:
                            upd_p = {"saldo_devedor": novo_saldo_p, "parcela_juros": novo_juros_p}
                            if novo_saldo_p == 0:
                                upd_p["status"] = "quitado"
                            sb.table("emprestimos_concedidos").update(upd_p).eq("id", emp_id_p).execute()
                        load_emprestimos_concedidos.clear()
                        load_pagamentos_recebidos.clear()
                        if eh_somente_juros:
                            st.success(
                                f"✅ Recebimento registrado! Contrato: **{sel_pag}** · "
                                f"Somente juros: **{brl(valor_p)}** · Saldo devedor não alterado."
                            )
                        else:
                            st.success(
                                f"✅ Recebimento registrado! "
                                f"Contrato: **{sel_pag}** · "
                                f"Amort.: **{brl(amort_p)}** · "
                                f"Novo saldo: **{brl(novo_saldo_p)}**"
                                + (" 🎉 Quitado!" if novo_saldo_p == 0 else "")
                            )
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao registrar: {e}")

    # ── Histórico por contrato ─────────────────────────────────────────────────
    # Mantém o expander aberto se há formulário de edição/exclusão ativo,
    # ou se acabou de chegar via link HTML (query param _open_dev).
    _hist_has_active = (not pagtos_d.empty) and any(
        st.session_state.get(f"editing_pag_{int(r['id'])}")
        or st.session_state.get(f"confirm_del_pag_{int(r['id'])}")
        for _, r in pagtos_d.iterrows()
    )
    _hist_force_open = (st.session_state.pop("_open_dev", None) == dev_id)
    with st.expander("📜 Histórico de Pagamentos por Contrato",
                     expanded=_hist_has_active or _hist_force_open):
        if pagtos_d.empty:
            st.info("Nenhum pagamento registrado ainda.")
        else:
            todos = pd.concat([ativos_d, quit_d], ignore_index=True) \
                    if not quit_d.empty else ativos_d.copy()
            for _, row_c in todos.sort_values("data_emprestimo").iterrows():
                eid      = int(row_c["id"])
                pagtos_c = pagtos_d[pagtos_d["emprestimo_id"] == eid].copy()
                if pagtos_c.empty:
                    continue
                pagtos_c = pagtos_c.sort_values(["data_pagamento", "id"], ascending=[False, False])
                badge    = "✅" if row_c["status"] == "quitado" else "📋"
                val_orig = float(row_c["valor_original"])
                taxa_c   = float(row_c["taxa_juros"])

                # ── Container com borda por contrato ───────────────────────
                with st.container(border=True):
                    st.markdown(
                        f"<b>{badge} {row_c['titulo']}</b>"
                        f"<span style='color:#666;font-size:0.85em'>"
                        f" &nbsp;·&nbsp; {len(pagtos_c)} pagamento(s)"
                        f" &nbsp;·&nbsp; Total pago: <b>{brl_md(float(pagtos_c['valor_pago'].sum()))}</b>"
                        f" &nbsp;·&nbsp; Saldo atual: <b>{brl_md(float(row_c['saldo_devedor']))}</b>"
                        f"</span>",
                        unsafe_allow_html=True
                    )
                    st.markdown("")

                    # ── Cabeçalho das colunas (HTML grid — alinha exato com as linhas) ──
                    _GCOLS = "1.4fr 1.4fr 1.1fr 1.4fr 1.5fr 2.0fr 0.85fr 0.85fr"
                    st.markdown(
                        f'<div style="display:grid;grid-template-columns:{_GCOLS};'
                        f'gap:6px;padding:4px 4px 2px 4px;font-size:0.82em;font-weight:700;">'
                        f'<span>Data de Pagamento</span><span>Valor Pago</span>'
                        f'<span>Juros</span><span>Amortização</span>'
                        f'<span>Saldo Após</span><span>Obs</span>'
                        f'<span></span><span></span></div>',
                        unsafe_allow_html=True,
                    )
                    st.divider()

                    # Contratos quitados: sem botões.
                    # Demais: só o lançamento do topo (data mais recente, desempate por id) tem botões.
                    # Como pagtos_c já está sorted desc por [data_pagamento, id], iloc[0] é o mais recente.
                    is_quitado_c  = (row_c["status"] == "quitado") or (float(row_c.get("saldo_devedor", 1)) <= 0)
                    ultimo_pag_id = -1 if is_quitado_c else int(pagtos_c.iloc[0]["id"])

                    # Saldo Após retroativo: computa on-the-fly onde saldo_depois é nulo (importados).
                    # Apenas para contratos ativos — quitados mantêm "—" conforme solicitado.
                    _saldo_retro = {}
                    if not is_quitado_c:
                        _asc = pagtos_c.sort_values(["data_pagamento", "id"], ascending=[True, True])
                        _cum = 0.0
                        for _, _p in _asc.iterrows():
                            _cum += float(_p.get("amortizacao") or 0)
                            if pd.isna(_p["saldo_depois"]):  # cobre None e NaN — coluna vem como dtype object quando só há NULLs
                                _saldo_retro[int(_p["id"])] = round(max(0.0, val_orig - _cum), 2)

                    for _, pr in pagtos_c.iterrows():
                        pag_id    = int(pr["id"])
                        ed_key    = f"editing_pag_{pag_id}"
                        del_key   = f"confirm_del_pag_{pag_id}"
                        is_ultimo = (pag_id == ultimo_pag_id)

                        # ── Chaves de session_state para os campos de edição ──
                        _sk = {
                            "data":  f"sv_data_{pag_id}",
                            "val":   f"sv_val_{pag_id}",
                            "obs":   f"sv_obs_{pag_id}",
                            "juros": f"sv_juros_{pag_id}",
                        }

                        # ── Estados mutuamente exclusivos: edit > delete > normal ──
                        if st.session_state.get(ed_key):
                            # ── Edição sem st.form (mais confiável em loops aninhados) ──
                            # Amortização não é campo editável: amort = valor_pago (amortização pura)
                            fe1, fe2, fe3, fe4 = st.columns(4)
                            # min_value dinâmico: se o pagamento for mais antigo que 365 dias,
                            # usa a data do próprio pagamento como mínimo (evita StreamlitAPIException)
                            _pag_dt = pr["data_pagamento"]
                            _pag_date = _pag_dt.date() if hasattr(_pag_dt, "date") else _pag_dt
                            _min_date = min(date.today() - timedelta(days=365*10), _pag_date)
                            fe1.date_input("Data", key=_sk["data"], format="DD/MM/YYYY",
                                           min_value=_min_date,
                                           max_value=date.today())
                            fe2.text_input("Valor pago (R$)", key=_sk["val"])
                            fe3.text_input("Juros ref. (R$)", key=_sk["juros"])
                            fe4.text_input("Observação", key=_sk["obs"])
                            fc1, fc2 = st.columns(2)

                            if fc1.button("💾 Salvar", key=f"btn_save_{pag_id}",
                                          type="primary", use_container_width=True):
                                _data_edit = st.session_state.get(_sk["data"])
                                _hoje_e    = date.today()
                                # Validação: apenas datas futuras são bloqueadas.
                                # Datas no passado são permitidas (consistente com min_value do date_input).
                                if _data_edit and _data_edit > _hoje_e:
                                    st.error("❌ Data futura não permitida.")
                                else:
                                    try:
                                        novo_val_pago  = round(parse_brl(st.session_state[_sk["val"]]), 2)
                                        _eh_juros_edit = (pr.get("tipo", "amortizacao") == "juros")
                                        # Pagamento "somente juros" nunca abate o saldo (amort = 0),
                                        # mesmo após edição — preserva o tipo original do lançamento.
                                        nova_amort = 0.0 if _eh_juros_edit else novo_val_pago
                                        # 1. Atualiza o registro do pagamento
                                        sb.table("pagamentos_recebidos").update({
                                            "data_pagamento": str(st.session_state[_sk["data"]]),
                                            "valor_pago":     novo_val_pago,
                                            "juros":          round(parse_brl(st.session_state[_sk["juros"]]), 2),
                                            "amortizacao":    nova_amort,
                                            "observacao":     st.session_state[_sk["obs"]],
                                        }).eq("id", pag_id).execute()
                                        # 2. Recalcula saldo do contrato a partir de todas as amortizações
                                        res_amort = sb.table("pagamentos_recebidos")\
                                            .select("amortizacao")\
                                            .eq("emprestimo_id", eid).execute()
                                        total_amort = sum(float(p["amortizacao"] or 0) for p in res_amort.data)
                                        novo_saldo  = round(max(0.0, val_orig - total_amort), 2)
                                        novo_juros  = round(novo_saldo * taxa_c, 2)
                                        # 3. Atualiza saldo_depois do pagamento = novo saldo do contrato
                                        sb.table("pagamentos_recebidos").update({
                                            "saldo_depois": novo_saldo,
                                        }).eq("id", pag_id).execute()
                                        # 4. Atualiza contrato
                                        upd_c = {
                                            "saldo_devedor": novo_saldo,
                                            "parcela_juros": novo_juros,
                                            "status": "quitado" if novo_saldo <= 0 else "ativo",
                                        }
                                        sb.table("emprestimos_concedidos").update(upd_c).eq("id", eid).execute()
                                        # 5. Limpa estado e recarrega
                                        for k in [ed_key] + list(_sk.values()):
                                            st.session_state.pop(k, None)
                                        load_emprestimos_concedidos.clear()
                                        load_pagamentos_recebidos.clear()
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Erro ao salvar: {e}")

                            if fc2.button("❌ Cancelar", key=f"btn_canc_{pag_id}",
                                          use_container_width=True):
                                for k in [ed_key] + list(_sk.values()):
                                    st.session_state.pop(k, None)
                                st.rerun()

                        elif st.session_state.get(del_key):
                            # ── Confirmação de exclusão ─────────────────────
                            st.warning(
                                f"Excluir pagamento de **{brl(pr['valor_pago'])}** "
                                f"em {pr['data_pagamento'].strftime('%d/%m/%Y')}?"
                            )
                            dc1, dc2 = st.columns(2)
                            if dc1.button("✅ Confirmar exclusão", key=f"conf_del_{pag_id}", type="primary"):
                                try:
                                    sb.table("pagamentos_recebidos").delete().eq("id", pag_id).execute()
                                    st.session_state.pop(del_key, None)
                                    # ── Recalcular saldo do contrato ────────
                                    res_amort = sb.table("pagamentos_recebidos")\
                                        .select("amortizacao")\
                                        .eq("emprestimo_id", eid).execute()
                                    total_amort = sum(float(p["amortizacao"] or 0) for p in res_amort.data)
                                    novo_saldo  = round(max(0.0, val_orig - total_amort), 2)
                                    novo_juros  = round(novo_saldo * taxa_c, 2)
                                    upd_c = {
                                        "saldo_devedor": novo_saldo,
                                        "parcela_juros": novo_juros,
                                        "status": "quitado" if novo_saldo <= 0 else "ativo",
                                    }
                                    sb.table("emprestimos_concedidos").update(upd_c).eq("id", eid).execute()
                                    load_emprestimos_concedidos.clear()
                                    load_pagamentos_recebidos.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Erro ao excluir: {e}")
                            if dc2.button("❌ Cancelar", key=f"canc_del_{pag_id}"):
                                st.session_state.pop(del_key, None)
                                st.rerun()

                        else:
                            # ── Linha de dados normal ───────────────────────
                            if is_quitado_c:
                                # Contrato quitado: linha com marca d'água verde, sem botões
                                _d   = pr["data_pagamento"].strftime("%d/%m/%Y")
                                _obs = str(pr["observacao"] or "")
                                if pr.get("tipo", "amortizacao") == "juros":
                                    _obs = "🔵 Somente Juros" + (f" · {_obs}" if _obs else "")
                                st.markdown(
                                    f"""<div style="
                                        display:grid;
                                        grid-template-columns:1.4fr 1.4fr 1.2fr 1.4fr 1.6fr 2.4fr 0.7fr;
                                        gap:6px;
                                        background:rgba(40,167,69,0.08);
                                        border-radius:6px;
                                        padding:6px 10px;
                                        align-items:center;
                                        margin-bottom:4px;
                                        font-size:0.88em;
                                    ">
                                        <span>{_d}</span>
                                        <span>{brl(pr['valor_pago'])}</span>
                                        <span>{brl(pr['juros'])}</span>
                                        <span>{brl(pr['amortizacao'])}</span>
                                        <span>{brl(pr['saldo_depois'])}</span>
                                        <span style="color:#666">{_obs}</span>
                                        <span style="
                                            color:#28a745;font-weight:700;
                                            font-size:0.75em;letter-spacing:0.08em;
                                            opacity:0.65;text-align:right;
                                        ">✓&nbsp;QUITADO</span>
                                    </div>""",
                                    unsafe_allow_html=True,
                                )
                            else:
                                # Contrato ativo: st.columns + st.button() (sem page reload)
                                # CSS :has() com IDs únicos por pag_id colore SOMENTE aquele botão.
                                # Também colapsa o stMarkdown da span a zero-height para alinhar.
                                # Usa "stColumn" (Streamlit 1.28+) E "column" (versões antigas) como fallback.
                                rc = st.columns([1.4, 1.4, 1.1, 1.4, 1.5, 2.0, 0.85, 0.85])
                                _sd = _saldo_retro.get(pag_id, pr["saldo_depois"])
                                rc[0].write(pr["data_pagamento"].strftime("%d/%m/%Y"))
                                rc[1].write(brl(pr["valor_pago"]))
                                rc[2].write(brl(pr["juros"]))
                                rc[3].write(brl(pr["amortizacao"]))
                                rc[4].write(brl(_sd))
                                _obs_c = str(pr["observacao"] or "")
                                if pr.get("tipo", "amortizacao") == "juros":
                                    _obs_c = "🔵 Somente Juros" + (f" · {_obs_c}" if _obs_c else "")
                                rc[5].write(_obs_c)
                                if is_ultimo:
                                    # ── EDITAR — azul ──────────────────────────
                                    # ORDEM: button PRIMEIRO, markdown depois.
                                    # Assim o botão fica no topo da coluna (alinhado com o texto),
                                    # e o container da span — colapsado a 0 — aparece abaixo sem impacto visual.
                                    if rc[6].button("EDITAR", key=f"btn_ed_{pag_id}",
                                                    use_container_width=True):
                                        dt = pr["data_pagamento"]
                                        st.session_state[_sk["data"]]  = dt.date() if hasattr(dt, "date") else dt
                                        st.session_state[_sk["val"]]   = brl_input(float(pr["valor_pago"] or 0))
                                        st.session_state[_sk["obs"]]   = str(pr["observacao"] or "")
                                        st.session_state[_sk["juros"]] = brl_input(float(pr["juros"] or 0))
                                        st.session_state[ed_key] = True
                                        st.rerun()
                                    rc[6].markdown(
                                        f'<style>'
                                        f'div[data-testid="stColumn"]:has(#em{pag_id}) button,'
                                        f'div[data-testid="column"]:has(#em{pag_id}) button'
                                        f'{{background:#1d4ed8!important;color:#fff!important;'
                                        f'border:none!important;border-radius:8px!important;'
                                        f'font-weight:700!important;font-size:.72rem!important;'
                                        f'letter-spacing:.06em!important;}}'
                                        f'div[data-testid="stColumn"]:has(#em{pag_id}) button:hover,'
                                        f'div[data-testid="column"]:has(#em{pag_id}) button:hover'
                                        f'{{background:#1e40af!important;'
                                        f'box-shadow:0 0 0 3px rgba(59,130,246,.4)!important;}}'
                                        # Colapsa o container da span a zero (fica abaixo do botão, sem espaço)
                                        f'div[data-testid="stColumn"]:has(#em{pag_id}) [data-testid="stMarkdown"],'
                                        f'div[data-testid="column"]:has(#em{pag_id}) [data-testid="stMarkdown"]'
                                        f'{{height:0!important;overflow:hidden!important;'
                                        f'min-height:0!important;margin:0!important;padding:0!important;}}'
                                        f'</style>'
                                        f'<span id="em{pag_id}"></span>',
                                        unsafe_allow_html=True,
                                    )
                                    # ── EXCLUIR — vermelho ─────────────────────
                                    if rc[7].button("EXCLUIR", key=f"btn_del_{pag_id}",
                                                    use_container_width=True):
                                        st.session_state[del_key] = True
                                        st.rerun()
                                    rc[7].markdown(
                                        f'<style>'
                                        f'div[data-testid="stColumn"]:has(#dm{pag_id}) button,'
                                        f'div[data-testid="column"]:has(#dm{pag_id}) button'
                                        f'{{background:#dc2626!important;color:#fff!important;'
                                        f'border:none!important;border-radius:8px!important;'
                                        f'font-weight:700!important;font-size:.72rem!important;'
                                        f'letter-spacing:.06em!important;}}'
                                        f'div[data-testid="stColumn"]:has(#dm{pag_id}) button:hover,'
                                        f'div[data-testid="column"]:has(#dm{pag_id}) button:hover'
                                        f'{{background:#b91c1c!important;'
                                        f'box-shadow:0 0 0 3px rgba(239,68,68,.4)!important;}}'
                                        f'div[data-testid="stColumn"]:has(#dm{pag_id}) [data-testid="stMarkdown"],'
                                        f'div[data-testid="column"]:has(#dm{pag_id}) [data-testid="stMarkdown"]'
                                        f'{{height:0!important;overflow:hidden!important;'
                                        f'min-height:0!important;margin:0!important;padding:0!important;}}'
                                        f'</style>'
                                        f'<span id="dm{pag_id}"></span>',
                                        unsafe_allow_html=True,
                                    )

                st.markdown("")

    # ── Novo Empréstimo ────────────────────────────────────────────────────────
    with st.expander(f"➕ Novo Empréstimo para {dev_nome}"):
        with st.form(key=f"form_ne_{dev_id}", clear_on_submit=True):
            nc1, nc2 = st.columns(2)
            new_titulo = nc1.text_input("Título")
            new_data   = nc2.date_input("Data do empréstimo",
                                         value=agora_br().date(),
                                         min_value=date.today() - timedelta(days=365),
                                         max_value=date.today(),
                                         format="DD/MM/YYYY")
            new_valor_str = nc1.text_input("Valor (R$)", value="0,00", placeholder="ex: 10.000,00",
                                           help="Digite o valor no formato 10.000,00")
            new_valor = parse_brl(new_valor_str)
            new_taxa_pct = nc2.number_input("Taxa a.m. (%)", min_value=0.0,
                                             max_value=10.0, step=0.01, format="%.2f")
            new_dia    = nc1.number_input("Dia de vencimento", min_value=1,
                                           max_value=31, step=1, value=10)
            submitted_ne = st.form_submit_button("💾 Cadastrar Empréstimo",
                                                  use_container_width=True)
        if submitted_ne:
            _hoje_ne = date.today()
            if new_data > _hoje_ne:
                st.error("❌ Data futura não permitida. Utilize uma data até hoje.")
            elif new_data < _hoje_ne - timedelta(days=365):
                st.error("❌ Data muito antiga. O limite máximo é 1 ano no passado.")
            elif not new_titulo or new_valor <= 0:
                st.warning("Preencha título e valor.")
            else:
                new_taxa    = new_taxa_pct / 100
                new_parcela = round(new_valor * new_taxa, 2)
                try:
                    sb.table("emprestimos_concedidos").insert({
                        "devedor_id":      dev_id,
                        "titulo":          new_titulo,
                        "data_emprestimo": str(new_data),
                        "valor_original":  round(new_valor, 2),
                        "saldo_devedor":   round(new_valor, 2),
                        "taxa_juros":      round(new_taxa, 6),
                        "parcela_juros":   new_parcela,
                        "dia_vencimento":  int(new_dia),
                        "status":          "ativo",
                    }).execute()
                    st.success(f"Empréstimo '{new_titulo}' cadastrado · "
                               f"Juros/mês: **{brl(new_parcela)}**")
                    load_emprestimos_concedidos.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro: {e}")

    # ── Excluir Contrato ───────────────────────────────────────────────────────
    todos_contratos = pd.concat([ativos_d, quit_d], ignore_index=True) \
                       if not quit_d.empty else ativos_d.copy()
    if not todos_contratos.empty:
        with st.expander("🗑️ Excluir Contrato de Empréstimo"):
            opcoes_del = todos_contratos["titulo"].tolist()
            sel_del  = st.selectbox("Contrato", opcoes_del, key=f"del_emp_sel_{dev_id}")
            row_del  = todos_contratos[todos_contratos["titulo"] == sel_del].iloc[0]
            eid_del  = int(row_del["id"])
            del_key  = f"confirm_del_emp_{eid_del}"
            n_pag_del = int((pagtos_d["emprestimo_id"] == eid_del).sum()) if not pagtos_d.empty else 0
            st.caption(
                f"Valor original: **{brl(float(row_del['valor_original']))}** · "
                f"Saldo devedor: **{brl(float(row_del['saldo_devedor']))}** · "
                f"{n_pag_del} pagamento(s) registrado(s)"
            )
            if st.session_state.get(del_key):
                st.warning(
                    f"Excluir o contrato **{sel_del}** definitivamente? "
                    + (f"Isso também apaga o(s) {n_pag_del} pagamento(s) registrado(s) nele. "
                       if n_pag_del else "")
                    + "Essa ação não pode ser desfeita."
                )
                dc1, dc2 = st.columns(2)
                if dc1.button("✅ Confirmar exclusão", key=f"conf_{del_key}",
                              type="primary", use_container_width=True):
                    try:
                        sb.table("emprestimos_concedidos").delete().eq("id", eid_del).execute()
                        st.session_state.pop(del_key, None)
                        load_emprestimos_concedidos.clear()
                        load_pagamentos_recebidos.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao excluir: {e}")
                if dc2.button("❌ Cancelar", key=f"canc_{del_key}", use_container_width=True):
                    st.session_state.pop(del_key, None)
                    st.rerun()
            else:
                if st.button("🗑️ Excluir Contrato", key=f"btn_{del_key}",
                              use_container_width=True):
                    st.session_state[del_key] = True
                    st.rerun()


# ─── TAB: EMPRÉSTIMOS ─────────────────────────────────────────────────────────
def tab_emprestimos(emp: pd.DataFrame):
    """Aba principal de empréstimos: concedidos (créditos) + tomados (débitos)."""
    # sub-navegação via session_state (não perde estado em reruns)
    _EMP_SUBS = ["🏦 Empréstimos Concedidos", "💳 Meus Empréstimos"]
    if "emp_sub_nav" not in st.session_state:
        st.session_state["emp_sub_nav"] = _EMP_SUBS[0]
    emp_nav = st.radio(
        "Sub-seção", _EMP_SUBS, horizontal=True,
        key="emp_sub_nav", label_visibility="collapsed"
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SUB-ABA 1: EMPRÉSTIMOS CONCEDIDOS (Lucas é o credor)
    # ══════════════════════════════════════════════════════════════════════════
    if emp_nav == "🏦 Empréstimos Concedidos":
        devedores_df  = load_devedores()
        emp_conc_df   = load_emprestimos_concedidos()
        pagtos_all    = load_pagamentos_recebidos()

        # ── KPIs globais ──────────────────────────────────────────────────────
        ativos_glob  = emp_conc_df[emp_conc_df["status"] == "ativo"] if not emp_conc_df.empty else pd.DataFrame()
        total_saldo  = float(ativos_glob["saldo_devedor"].sum()) if not ativos_glob.empty else 0.0
        total_juros  = float(ativos_glob["parcela_juros"].sum()) if not ativos_glob.empty else 0.0
        n_contratos  = len(ativos_glob)
        total_receb  = float(pagtos_all["valor_pago"].sum())    if not pagtos_all.empty else 0.0

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f"""<div class='kpi-card'><div class='kpi-value'>{brl(total_saldo)}</div>
            <div class='kpi-label'>💰 Total a Receber</div></div>""", unsafe_allow_html=True)
        c2.markdown(f"""<div class='kpi-card ouro'><div class='kpi-value'>{brl(total_juros)}</div>
            <div class='kpi-label'>📅 Juros Mensais</div></div>""", unsafe_allow_html=True)
        c3.markdown(f"""<div class='kpi-card'><div class='kpi-value'>{n_contratos}</div>
            <div class='kpi-label'>📋 Contratos Ativos</div></div>""", unsafe_allow_html=True)
        c4.markdown(f"""<div class='kpi-card'><div class='kpi-value'>{brl(total_receb)}</div>
            <div class='kpi-label'>✅ Total Recebido</div></div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════
        # IMPORTAR PLANILHA DE EMPRÉSTIMOS (.xlsx no padrão "Títulos")
        # ══════════════════════════════════════════════════════════════════════
        _import_plan_pending_msg = st.session_state.get("import_plan_msg")
        with st.expander("📥 Importar Planilha de Empréstimos (.xlsx)",
                          expanded=bool(_import_plan_pending_msg)):
            if _import_plan_pending_msg:
                st.session_state.pop("import_plan_msg", None)
                st.success(_import_plan_pending_msg)

            st.caption(
                "Sobe uma planilha no padrão \"Títulos\" (colunas Títulos, Data Emp., "
                "Valor Originário, Saldo Devedor, Parcela Juros...) e mostra uma prévia "
                "para conferência antes de gravar no banco."
            )
            imp_file = st.file_uploader("Planilha (.xlsx)", type=["xlsx"], key="import_plan_upload")

            if imp_file is not None:
                _file_id = f"{imp_file.name}_{imp_file.size}"
                if st.session_state.get("import_plan_file_id") != _file_id:
                    try:
                        st.session_state["import_plan_data"] = parse_planilha_titulos(imp_file.getvalue())
                    except Exception as e:
                        st.session_state["import_plan_data"] = {}
                        st.error(f"Erro ao ler a planilha: {e}")
                    st.session_state["import_plan_file_id"] = _file_id

                _parsed = st.session_state.get("import_plan_data", {})
                if not _parsed:
                    st.warning(
                        "Nenhuma aba no formato reconhecido foi encontrada nesta planilha. "
                        "Use o importador de planilhas para carregar contratos."
                    )
                else:
                    sel_sheet = st.selectbox(
                        "Aba detectada", list(_parsed.keys()), key="import_plan_sheet_sel")
                    _dados = _parsed[sel_sheet]

                    st.markdown(f"**{len(_dados['emprestimos'])} empréstimo(s)** encontrados nesta aba. "
                                "Revise/edite antes de importar:")
                    df_emp_edit = st.data_editor(
                        pd.DataFrame(_dados["emprestimos"]),
                        key=f"import_emp_editor_{sel_sheet}",
                        use_container_width=True, num_rows="dynamic", hide_index=True,
                    )

                    _titulos_atuais = [t for t in df_emp_edit["titulo"].tolist() if t]
                    matched, sem_match = match_pagamentos_titulos(_titulos_atuais, _dados["pagamentos_raw"])
                    _importaveis = [p for p in matched if p["data_pagamento"]]
                    _avisos      = sem_match + [p for p in matched if not p["data_pagamento"]]

                    st.markdown(
                        f"**{len(_importaveis)} pagamento(s)** conciliados e prontos para importar"
                        + (f" · **{len(_avisos)} sem correspondência/data** (não serão importados)"
                           if _avisos else "") + "."
                    )
                    if _avisos:
                        with st.expander(f"⚠️ {len(_avisos)} pagamento(s) que ficarão de fora"):
                            st.dataframe(pd.DataFrame(_avisos), use_container_width=True, hide_index=True)

                    st.markdown("---")
                    _nomes_dev = devedores_df["nome"].tolist() if not devedores_df.empty else []
                    _opts_dev  = ["➕ Novo devedor"] + _nomes_dev
                    _sugestao  = _dados.get("devedor_sugerido", "")
                    _default_idx = _opts_dev.index(_sugestao) if _sugestao in _nomes_dev else 0
                    sel_dev_imp = st.selectbox("Devedor de destino", _opts_dev,
                                                index=_default_idx, key="import_plan_devedor_sel")
                    novo_nome_dev = ""
                    if sel_dev_imp == "➕ Novo devedor":
                        novo_nome_dev = st.text_input(
                            "Nome do novo devedor", value=_sugestao, key="import_plan_devedor_novo")

                    if st.button("✅ Confirmar Importação", type="primary",
                                 use_container_width=True, key="btn_confirmar_import_plan"):
                        _nome_final = novo_nome_dev.strip() if sel_dev_imp == "➕ Novo devedor" else sel_dev_imp
                        if not _nome_final:
                            st.error("Informe o nome do devedor.")
                        elif df_emp_edit.empty:
                            st.error("Nenhum empréstimo para importar.")
                        else:
                            try:
                                # Reaproveita devedor existente com o mesmo nome (case/espaço-insensitive)
                                # mesmo que o usuário tenha marcado "Novo devedor" — evita duplicatas.
                                _match_existente = devedores_df[
                                    devedores_df["nome"].str.strip().str.lower() == _nome_final.lower()
                                ] if not devedores_df.empty else pd.DataFrame()
                                if not _match_existente.empty:
                                    dev_id_imp = int(_match_existente.iloc[0]["id"])
                                elif sel_dev_imp == "➕ Novo devedor":
                                    res_dev = sb.table("devedores").insert({
                                        "nome": _nome_final, "categoria": "", "ativo": True,
                                    }).execute()
                                    dev_id_imp = res_dev.data[0]["id"]
                                else:
                                    dev_id_imp = int(
                                        devedores_df[devedores_df["nome"] == _nome_final].iloc[0]["id"])

                                emp_id_map = {}
                                for _, er in df_emp_edit.iterrows():
                                    if not er.get("titulo"):
                                        continue
                                    _valor   = float(er["valor_original"] or 0)
                                    _saldo   = float(er["saldo_devedor"] or 0)
                                    _parcela = float(er["parcela_juros"] or 0)
                                    _taxa    = round(_parcela / _valor, 6) if _valor > 0 and _parcela > 0 else 0.02
                                    _dv      = er.get("dia_vencimento")
                                    _de      = er.get("data_emprestimo")
                                    res_e = sb.table("emprestimos_concedidos").insert({
                                        "devedor_id":      dev_id_imp,
                                        "titulo":          er["titulo"],
                                        "data_emprestimo": str(_de) if _de and not pd.isna(_de) else None,
                                        "valor_original":  round(_valor, 2),
                                        "saldo_devedor":   round(_saldo, 2),
                                        "taxa_juros":      _taxa,
                                        "parcela_juros":   round(_parcela, 2),
                                        "recorrencia":     er.get("recorrencia") or "Mensal",
                                        "dia_vencimento":  int(_dv) if _dv and not pd.isna(_dv) else None,
                                        "status":          "ativo" if _saldo > 0 else "quitado",
                                    }).execute()
                                    emp_id_map[er["titulo"]] = res_e.data[0]["id"]

                                _n_pag_ok = 0
                                for pg in _importaveis:
                                    _eid = emp_id_map.get(pg["titulo_emprestimo"])
                                    if _eid is None:
                                        continue
                                    sb.table("pagamentos_recebidos").insert({
                                        "emprestimo_id":  _eid,
                                        "data_pagamento": str(pg["data_pagamento"]),
                                        "valor_pago":     round(float(pg["valor_pago"] or 0), 2),
                                        "juros":          0,
                                        "amortizacao":    round(float(pg["valor_pago"] or 0), 2),
                                        "observacao":     "Importado da planilha",
                                    }).execute()
                                    _n_pag_ok += 1

                                load_devedores.clear()
                                load_emprestimos_concedidos.clear()
                                load_pagamentos_recebidos.clear()
                                # Limpa TAMBÉM o file_uploader (senão o Streamlit mantém o
                                # arquivo "preso" no widget e ele é reprocessado/reimportado
                                # nos reruns seguintes, criando devedores duplicados).
                                for _k in ("import_plan_file_id", "import_plan_data",
                                           "import_plan_upload", "import_plan_sheet_sel",
                                           "import_plan_devedor_sel", "import_plan_devedor_novo"):
                                    st.session_state.pop(_k, None)
                                st.session_state["import_plan_msg"] = (
                                    f"Importado! {len(emp_id_map)} empréstimo(s) e "
                                    f"{_n_pag_ok} pagamento(s) para **{_nome_final}**."
                                )
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro ao importar: {e}")

        if devedores_df.empty:
            st.info("Nenhum devedor cadastrado ainda.")
            with st.form("form_novo_devedor_first"):
                nd_nome = st.text_input("Nome do devedor")
                nd_cat  = st.text_input("Categoria (ex: Loja, Pessoa)")
                if st.form_submit_button("💾 Cadastrar"):
                    if nd_nome:
                        try:
                            sb.table("devedores").insert({"nome": nd_nome, "categoria": nd_cat, "ativo": True}).execute()
                            load_devedores.clear()
                            st.success("Devedor cadastrado!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro: {e}")
        else:
            # ── Seletor de devedor (selectbox em vez de tabs aninhadas) ───────
            nomes_dev = [row["nome"] for _, row in devedores_df.iterrows()]
            opcoes    = nomes_dev + ["➕ Cadastrar Novo Devedor"]
            sel_dev   = st.selectbox("Devedor", opcoes, key="sel_devedor_conc",
                                     label_visibility="collapsed")

            if sel_dev == "➕ Cadastrar Novo Devedor":
                st.markdown("### Novo Devedor")
                with st.form("form_novo_devedor"):
                    nd1, nd2 = st.columns(2)
                    nd_nome     = nd1.text_input("Nome")
                    nd_categoria= nd2.text_input("Categoria (ex: Loja, Pessoa)")
                    nd_contato  = st.text_input("Contato (opcional)")
                    if st.form_submit_button("💾 Cadastrar Devedor", use_container_width=True):
                        if not nd_nome:
                            st.warning("Informe o nome.")
                        else:
                            try:
                                sb.table("devedores").insert({
                                    "nome": nd_nome, "categoria": nd_categoria,
                                    "contato": nd_contato, "ativo": True,
                                }).execute()
                                st.success(f"Devedor '{nd_nome}' cadastrado!")
                                load_devedores.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro: {e}")
            else:
                dev_row  = devedores_df[devedores_df["nome"] == sel_dev].iloc[0]
                dev_id   = int(dev_row["id"])
                dev_nome = str(dev_row["nome"])
                _conteudo_devedor(dev_id, dev_nome)


    # ══════════════════════════════════════════════════════════════════════════
    # SUB-ABA 2: MEUS EMPRÉSTIMOS (Lucas é o devedor)
    # ══════════════════════════════════════════════════════════════════════════
    elif emp_nav == "💳 Meus Empréstimos":
        def _form_novo():
            with st.expander("➕ Cadastrar Novo Empréstimo", expanded=emp.empty):
                c1, c2 = st.columns(2)
                novo_credor  = c1.text_input("Credor", key="ne_credor")
                novo_titulo  = c2.text_input("Título / Descrição", key="ne_titulo")
                novo_saldo_str = c1.text_input("Saldo devedor (R$)", value="0,00",
                                               placeholder="ex: 50.000,00", key="ne_saldo",
                                               help="Digite o valor no formato 50.000,00")
                novo_saldo_v = parse_brl(novo_saldo_str)
                nova_taxa_v  = c2.number_input("Taxa a.m. (%)", min_value=0.0,
                                               max_value=100.0, step=0.01,
                                               format="%.2f", key="ne_taxa")
                nova_parcela = round(novo_saldo_v * nova_taxa_v / 100, 2)
                st.markdown(f'<small style="color:#555">Juros/mês estimados: <b>{brl(nova_parcela)}</b></small>', unsafe_allow_html=True)
                if st.button("💾 Cadastrar Empréstimo", key="btn_ne", use_container_width=True):
                    if not novo_credor or not novo_titulo or novo_saldo_v <= 0:
                        st.warning("Preencha todos os campos obrigatórios.")
                    else:
                        try:
                            sb.table("emprestimos").insert({
                                "credor": novo_credor, "titulo": novo_titulo,
                                "saldo_devedor": round(novo_saldo_v, 2),
                                "taxa_juros": round(nova_taxa_v / 100, 6),
                                "parcela_juros": nova_parcela, "status": "ativo",
                            }).execute()
                            st.success("Empréstimo cadastrado!")
                            load_emprestimos.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao cadastrar: {e}")

        if emp.empty:
            st.info("Nenhum empréstimo ativo cadastrado.")
            _form_novo()
        else:
            total_divida = float(emp["saldo_devedor"].sum())
            total_juros_me  = float(emp["parcela_juros"].sum())
            c1, c2, c3 = st.columns(3)
            c1.markdown(f"""<div class='kpi-card ouro'>
                <div class='kpi-value'>{brl(total_divida)}</div>
                <div class='kpi-label'>💳 Total de Dívidas</div>
            </div>""", unsafe_allow_html=True)
            c2.markdown(f"""<div class='kpi-card neg'>
                <div class='kpi-value'>{brl(total_juros_me)}</div>
                <div class='kpi-label'>🔴 Juros Mensais</div>
            </div>""", unsafe_allow_html=True)
            c3.markdown(f"""<div class='kpi-card'>
                <div class='kpi-value'>{len(emp)}</div>
                <div class='kpi-label'>📋 Contratos Ativos</div>
            </div>""", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            grp = emp.groupby("credor")["saldo_devedor"].sum().reset_index()
            col_g, col_t = st.columns([1, 1.8])
            with col_g:
                fig2 = go.Figure(go.Pie(
                    labels=grp["credor"], values=grp["saldo_devedor"],
                    hole=0.5, marker_colors=[VERDE, "#3A7D5A", OURO, "#D4A017"],
                    textinfo="label+percent", textfont_size=12,
                ))
                fig2.update_layout(height=260, margin=dict(l=0,r=0,t=0,b=0),
                                   showlegend=False, paper_bgcolor="#FFFBEF")
                st.plotly_chart(fig2, use_container_width=True)
            with col_t:
                df_show = emp[["credor","titulo","saldo_devedor","taxa_juros","parcela_juros"]].copy()
                df_show.columns = ["Credor","Título","Saldo Devedor","Taxa a.m.","Juros/Mês"]
                df_show["Saldo Devedor"] = df_show["Saldo Devedor"].apply(brl)
                df_show["Taxa a.m."]     = df_show["Taxa a.m."].apply(lambda x: f"{float(x)*100:.2f}%")
                df_show["Juros/Mês"]     = df_show["Juros/Mês"].apply(brl)
                st.dataframe(df_show, use_container_width=True, hide_index=True)

            st.markdown("---")
            with st.expander("💸 Registrar Pagamento"):
                contratos_me = emp["titulo"].tolist()
                sel_c        = st.selectbox("Contrato", contratos_me, key="pag_sel")
                row_c        = emp[emp["titulo"] == sel_c].iloc[0]
                saldo_at     = float(row_c["saldo_devedor"])
                juros_mes_v  = float(row_c["parcela_juros"])
                st.markdown(
                    f'<small style="color:#555">Saldo atual: <b>{brl_md(saldo_at)}</b> &nbsp;|&nbsp; Juros do mês: <b>{brl_md(juros_mes_v)}</b></small>',
                    unsafe_allow_html=True)
                col_vp, col_dp = st.columns(2)
                valor_pago_str = col_vp.text_input("Valor pago (R$)",
                                                    value=brl_input(juros_mes_v),
                                                    placeholder="ex: 800,00", key="pag_val",
                                                    help="Digite o valor no formato 800,00 ou 2.000,50")
                valor_pago_v = parse_brl(valor_pago_str)
                data_pag_v   = col_dp.date_input("Data do pagamento",
                                                  value=agora_br().date(),
                                                  min_value=date.today() - timedelta(days=365),
                                                  max_value=date.today(),
                                                  key="pag_data")
                obs_v        = st.text_input("Observação", key="pag_obs")
                juros_no_pag = min(valor_pago_v, juros_mes_v)
                amort_v      = max(0.0, valor_pago_v - juros_no_pag)
                novo_saldo_v = max(0.0, saldo_at - amort_v)
                st.caption(f"↳ Juros: **{brl(juros_no_pag)}** | Amortização: **{brl(amort_v)}** | Novo saldo: **{brl(novo_saldo_v)}**")
                if st.button("✅ Confirmar Pagamento", key="btn_pag",
                             use_container_width=True, type="primary"):
                    _hoje_v = date.today()
                    if data_pag_v > _hoje_v:
                        st.error("❌ Data futura não permitida. Utilize uma data até hoje.")
                    elif data_pag_v < _hoje_v - timedelta(days=365):
                        st.error("❌ Data muito antiga. O limite máximo é 1 ano no passado.")
                    else:
                        try:
                            emp_id = int(row_c["id"])
                            try:
                                res_h  = sb.table("emprestimos").select("historico_pagamentos").eq("id", emp_id).execute()
                                hist_v = (res_h.data[0].get("historico_pagamentos") or []) if res_h.data else []
                                hist_v.append({
                                    "data": str(data_pag_v), "valor_pago": round(valor_pago_v, 2),
                                    "juros": round(juros_no_pag, 2), "amortizacao": round(amort_v, 2),
                                    "saldo_antes": round(saldo_at, 2), "saldo_depois": round(novo_saldo_v, 2),
                                    "obs": obs_v,
                                })
                                upd = {"saldo_devedor": round(novo_saldo_v, 2),
                                       "parcela_juros": round(novo_saldo_v * float(row_c["taxa_juros"]), 2),
                                       "historico_pagamentos": hist_v}
                            except Exception:
                                upd = {"saldo_devedor": round(novo_saldo_v, 2),
                                       "parcela_juros": round(novo_saldo_v * float(row_c["taxa_juros"]), 2)}
                            if novo_saldo_v == 0:
                                upd["status"] = "quitado"
                            sb.table("emprestimos").update(upd).eq("id", emp_id).execute()
                            st.success(f"Pagamento registrado! Novo saldo: **{brl(novo_saldo_v)}**")
                            load_emprestimos.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro: {e}")

            with st.expander("✅ Marcar Contrato como Quitado"):
                sel_q = st.selectbox("Contrato", emp["titulo"].tolist(), key="quit_sel")
                row_q = emp[emp["titulo"] == sel_q].iloc[0]
                st.caption(f"Saldo a ser baixado: **{brl(float(row_q['saldo_devedor']))}**")
                st.warning("Essa ação marca o contrato como quitado e zera o saldo.")
                if st.button("🏁 Confirmar Quitação", key="btn_quit", use_container_width=True):
                    try:
                        sb.table("emprestimos").update(
                            {"status": "quitado", "saldo_devedor": 0.0, "parcela_juros": 0.0}
                        ).eq("id", int(row_q["id"])).execute()
                        st.success(f"Contrato '{sel_q}' quitado! 🎉")
                        load_emprestimos.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro: {e}")

            _form_novo()

            def _fmt_data_me(v):
                try:
                    return datetime.strptime(str(v), "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    return str(v or "—")

            _hist_has_active_me = any(
                v for k, v in st.session_state.items()
                if k.startswith("editing_me_") or k.startswith("confirm_del_me_")
            )
            with st.expander("📜 Histórico de Pagamentos por Contrato",
                             expanded=_hist_has_active_me):
                try:
                    todos_emp = sb.table("emprestimos").select(
                        "id,titulo,credor,status,saldo_devedor,taxa_juros,historico_pagamentos"
                    ).execute()
                except Exception as e:
                    st.error(f"⚠️ Erro ao carregar histórico. ({e})")
                    todos_emp = None

                tem_hist = False
                if todos_emp and todos_emp.data:
                    for r_h in todos_emp.data:
                        hist_rows = r_h.get("historico_pagamentos") or []
                        if not hist_rows:
                            continue
                        tem_hist = True
                        eid      = r_h["id"]
                        status_c = r_h.get("status", "ativo")
                        saldo_c  = float(r_h.get("saldo_devedor") or 0)
                        taxa_c   = float(r_h.get("taxa_juros") or 0)
                        is_quitado_c = (status_c == "quitado") or (saldo_c <= 0)

                        # Lista é append-only (novos pagamentos sempre no final),
                        # então o último índice é sempre o pagamento mais recente.
                        _indexed   = list(enumerate(hist_rows))[::-1]
                        ultimo_idx = -1 if is_quitado_c else (len(hist_rows) - 1)

                        with st.container(border=True):
                            badge_c      = "✅" if is_quitado_c else "📋"
                            total_pago_c = sum(float(h.get("valor_pago") or 0) for h in hist_rows)
                            st.markdown(
                                f"<b>{badge_c} {r_h['titulo']}</b>"
                                f"<span style='color:#666;font-size:0.85em'>"
                                f" &nbsp;·&nbsp; {r_h.get('credor','')}"
                                f" &nbsp;·&nbsp; {len(hist_rows)} pagamento(s)"
                                f" &nbsp;·&nbsp; Total pago: <b>{brl_md(total_pago_c)}</b>"
                                f" &nbsp;·&nbsp; Saldo atual: <b>{brl_md(saldo_c)}</b>"
                                f"</span>",
                                unsafe_allow_html=True,
                            )
                            st.markdown("")

                            _GCOLS_ME = "1.4fr 1.4fr 1.1fr 1.4fr 1.5fr 2.0fr 0.85fr 0.85fr"
                            st.markdown(
                                f'<div style="display:grid;grid-template-columns:{_GCOLS_ME};'
                                f'gap:6px;padding:4px 4px 2px 4px;font-size:0.82em;font-weight:700;">'
                                f'<span>Data de Pagamento</span><span>Valor Pago</span>'
                                f'<span>Juros</span><span>Amortização</span>'
                                f'<span>Saldo Após</span><span>Obs</span>'
                                f'<span></span><span></span></div>',
                                unsafe_allow_html=True,
                            )
                            st.divider()

                            for idx, hp in _indexed:
                                key       = f"{eid}_{idx}"
                                ed_key    = f"editing_me_{key}"
                                del_key   = f"confirm_del_me_{key}"
                                is_ultimo = (idx == ultimo_idx)

                                _sk = {
                                    "data":  f"sv_data_me_{key}",
                                    "val":   f"sv_val_me_{key}",
                                    "obs":   f"sv_obs_me_{key}",
                                    "juros": f"sv_juros_me_{key}",
                                }

                                # ── Estados mutuamente exclusivos: edit > delete > normal ──
                                if st.session_state.get(ed_key):
                                    fe1, fe2, fe3, fe4 = st.columns(4)
                                    try:
                                        _pag_date = datetime.strptime(str(hp.get("data")), "%Y-%m-%d").date()
                                    except Exception:
                                        _pag_date = date.today()
                                    _min_date = min(date.today() - timedelta(days=365*10), _pag_date)
                                    fe1.date_input("Data", key=_sk["data"], format="DD/MM/YYYY",
                                                   min_value=_min_date, max_value=date.today())
                                    fe2.text_input("Valor pago (R$)", key=_sk["val"])
                                    fe3.text_input("Juros ref. (R$)", key=_sk["juros"])
                                    fe4.text_input("Observação", key=_sk["obs"])
                                    fc1, fc2 = st.columns(2)

                                    if fc1.button("💾 Salvar", key=f"btn_save_me_{key}",
                                                  type="primary", use_container_width=True):
                                        _data_edit = st.session_state.get(_sk["data"])
                                        _hoje_e    = date.today()
                                        if _data_edit and _data_edit > _hoje_e:
                                            st.error("❌ Data futura não permitida.")
                                        else:
                                            try:
                                                novo_val_pago  = round(parse_brl(st.session_state[_sk["val"]]), 2)
                                                juros_digitado = round(parse_brl(st.session_state[_sk["juros"]]), 2)
                                                juros_aplicado = round(min(novo_val_pago, max(0.0, juros_digitado)), 2)
                                                nova_amort     = round(max(0.0, novo_val_pago - juros_aplicado), 2)
                                                saldo_antes_e  = float(hp.get("saldo_antes") or 0)
                                                novo_saldo_e   = round(max(0.0, saldo_antes_e - nova_amort), 2)

                                                # Relê o array atual (evita sobrescrever concorrência)
                                                res_cur = sb.table("emprestimos").select(
                                                    "historico_pagamentos").eq("id", eid).execute()
                                                hist_atual = (res_cur.data[0].get("historico_pagamentos") or []) \
                                                             if res_cur.data else []
                                                if idx < len(hist_atual):
                                                    hist_atual[idx] = {
                                                        "data":         str(_data_edit),
                                                        "valor_pago":   novo_val_pago,
                                                        "juros":        juros_aplicado,
                                                        "amortizacao":  nova_amort,
                                                        "saldo_antes":  round(saldo_antes_e, 2),
                                                        "saldo_depois": novo_saldo_e,
                                                        "obs":          st.session_state[_sk["obs"]],
                                                    }
                                                novo_parcela_e = round(novo_saldo_e * taxa_c, 2)
                                                upd_e = {
                                                    "historico_pagamentos": hist_atual,
                                                    "saldo_devedor": novo_saldo_e,
                                                    "parcela_juros": novo_parcela_e,
                                                    "status": "quitado" if novo_saldo_e <= 0 else "ativo",
                                                }
                                                sb.table("emprestimos").update(upd_e).eq("id", eid).execute()
                                                for k in [ed_key] + list(_sk.values()):
                                                    st.session_state.pop(k, None)
                                                load_emprestimos.clear()
                                                st.rerun()
                                            except Exception as e:
                                                st.error(f"Erro ao salvar: {e}")

                                    if fc2.button("❌ Cancelar", key=f"btn_canc_me_{key}",
                                                  use_container_width=True):
                                        for k in [ed_key] + list(_sk.values()):
                                            st.session_state.pop(k, None)
                                        st.rerun()

                                elif st.session_state.get(del_key):
                                    st.warning(
                                        f"Excluir pagamento de **{brl(hp.get('valor_pago'))}** "
                                        f"em {_fmt_data_me(hp.get('data'))}?"
                                    )
                                    dc1, dc2 = st.columns(2)
                                    if dc1.button("✅ Confirmar exclusão", key=f"conf_del_me_{key}", type="primary"):
                                        try:
                                            res_cur = sb.table("emprestimos").select(
                                                "historico_pagamentos").eq("id", eid).execute()
                                            hist_atual = (res_cur.data[0].get("historico_pagamentos") or []) \
                                                         if res_cur.data else []
                                            _removido = hist_atual.pop(idx) if idx < len(hist_atual) else hp
                                            novo_saldo_d   = round(float(_removido.get("saldo_antes") or 0), 2)
                                            novo_parcela_d = round(novo_saldo_d * taxa_c, 2)
                                            upd_d = {
                                                "historico_pagamentos": hist_atual,
                                                "saldo_devedor": novo_saldo_d,
                                                "parcela_juros": novo_parcela_d,
                                                "status": "quitado" if novo_saldo_d <= 0 else "ativo",
                                            }
                                            sb.table("emprestimos").update(upd_d).eq("id", eid).execute()
                                            st.session_state.pop(del_key, None)
                                            load_emprestimos.clear()
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Erro ao excluir: {e}")
                                    if dc2.button("❌ Cancelar", key=f"canc_del_me_{key}"):
                                        st.session_state.pop(del_key, None)
                                        st.rerun()

                                else:
                                    if is_quitado_c:
                                        _d   = _fmt_data_me(hp.get("data"))
                                        _obs = str(hp.get("obs") or "")
                                        st.markdown(
                                            f"""<div style="
                                                display:grid;
                                                grid-template-columns:1.4fr 1.4fr 1.2fr 1.4fr 1.6fr 2.4fr 0.7fr;
                                                gap:6px;
                                                background:rgba(40,167,69,0.08);
                                                border-radius:6px;
                                                padding:6px 10px;
                                                align-items:center;
                                                margin-bottom:4px;
                                                font-size:0.88em;
                                            ">
                                                <span>{_d}</span>
                                                <span>{brl(hp.get('valor_pago'))}</span>
                                                <span>{brl(hp.get('juros'))}</span>
                                                <span>{brl(hp.get('amortizacao'))}</span>
                                                <span>{brl(hp.get('saldo_depois'))}</span>
                                                <span style="color:#666">{_obs}</span>
                                                <span style="
                                                    color:#28a745;font-weight:700;
                                                    font-size:0.75em;letter-spacing:0.08em;
                                                    opacity:0.65;text-align:right;
                                                ">✓&nbsp;QUITADO</span>
                                            </div>""",
                                            unsafe_allow_html=True,
                                        )
                                    else:
                                        rc = st.columns([1.4, 1.4, 1.1, 1.4, 1.5, 2.0, 0.85, 0.85])
                                        rc[0].write(_fmt_data_me(hp.get("data")))
                                        rc[1].write(brl(hp.get("valor_pago")))
                                        rc[2].write(brl(hp.get("juros")))
                                        rc[3].write(brl(hp.get("amortizacao")))
                                        rc[4].write(brl(hp.get("saldo_depois")))
                                        rc[5].write(str(hp.get("obs") or ""))
                                        if is_ultimo:
                                            # ── EDITAR — azul ──────────────────────────
                                            if rc[6].button("EDITAR", key=f"btn_ed_me_{key}",
                                                            use_container_width=True):
                                                try:
                                                    _dt_me = datetime.strptime(
                                                        str(hp.get("data")), "%Y-%m-%d").date()
                                                except Exception:
                                                    _dt_me = date.today()
                                                st.session_state[_sk["data"]]  = _dt_me
                                                st.session_state[_sk["val"]]   = brl_input(float(hp.get("valor_pago") or 0))
                                                st.session_state[_sk["obs"]]   = str(hp.get("obs") or "")
                                                st.session_state[_sk["juros"]] = brl_input(float(hp.get("juros") or 0))
                                                st.session_state[ed_key] = True
                                                st.rerun()
                                            rc[6].markdown(
                                                f'<style>'
                                                f'div[data-testid="stColumn"]:has(#meEm{key}) button,'
                                                f'div[data-testid="column"]:has(#meEm{key}) button'
                                                f'{{background:#1d4ed8!important;color:#fff!important;'
                                                f'border:none!important;border-radius:8px!important;'
                                                f'font-weight:700!important;font-size:.72rem!important;'
                                                f'letter-spacing:.06em!important;}}'
                                                f'div[data-testid="stColumn"]:has(#meEm{key}) button:hover,'
                                                f'div[data-testid="column"]:has(#meEm{key}) button:hover'
                                                f'{{background:#1e40af!important;'
                                                f'box-shadow:0 0 0 3px rgba(59,130,246,.4)!important;}}'
                                                f'div[data-testid="stColumn"]:has(#meEm{key}) [data-testid="stMarkdown"],'
                                                f'div[data-testid="column"]:has(#meEm{key}) [data-testid="stMarkdown"]'
                                                f'{{height:0!important;overflow:hidden!important;'
                                                f'min-height:0!important;margin:0!important;padding:0!important;}}'
                                                f'</style>'
                                                f'<span id="meEm{key}"></span>',
                                                unsafe_allow_html=True,
                                            )
                                            # ── EXCLUIR — vermelho ─────────────────────
                                            if rc[7].button("EXCLUIR", key=f"btn_del_me_{key}",
                                                            use_container_width=True):
                                                st.session_state[del_key] = True
                                                st.rerun()
                                            rc[7].markdown(
                                                f'<style>'
                                                f'div[data-testid="stColumn"]:has(#meDm{key}) button,'
                                                f'div[data-testid="column"]:has(#meDm{key}) button'
                                                f'{{background:#dc2626!important;color:#fff!important;'
                                                f'border:none!important;border-radius:8px!important;'
                                                f'font-weight:700!important;font-size:.72rem!important;'
                                                f'letter-spacing:.06em!important;}}'
                                                f'div[data-testid="stColumn"]:has(#meDm{key}) button:hover,'
                                                f'div[data-testid="column"]:has(#meDm{key}) button:hover'
                                                f'{{background:#b91c1c!important;'
                                                f'box-shadow:0 0 0 3px rgba(239,68,68,.4)!important;}}'
                                                f'div[data-testid="stColumn"]:has(#meDm{key}) [data-testid="stMarkdown"],'
                                                f'div[data-testid="column"]:has(#meDm{key}) [data-testid="stMarkdown"]'
                                                f'{{height:0!important;overflow:hidden!important;'
                                                f'min-height:0!important;margin:0!important;padding:0!important;}}'
                                                f'</style>'
                                                f'<span id="meDm{key}"></span>',
                                                unsafe_allow_html=True,
                                            )
                        st.markdown("")

                if not tem_hist:
                    st.info("Nenhum pagamento via sistema ainda.")



# ─── TAB: INVESTIMENTOS ESCRITÓRIO ────────────────────────────────────────────
def tab_escritorio(inv: pd.DataFrame):
    # ── KPIs ─────────────────────────────────────────────────────────────────
    TAXA_ESCRITORIO = 0.01  # rendimento = sempre 1% do saldo do mês anterior

    def _form_novo_mes():
        with st.expander("➕ Lançar Novo Mês"):
            inv_plot2 = inv[inv["saldo_final"] > 0] if not inv.empty else inv
            ultimo_saldo = float(inv_plot2.iloc[-1]["saldo_final"]) if not inv_plot2.empty else 0.0
            c1m, c2m = st.columns(2)
            mes_v   = c1m.date_input("Mês de referência (dia 1)", key="em_mes",
                                      format="DD/MM/YYYY")
            tipo_v  = c2m.selectbox("Tipo", ["Honorários", "Êxito", "Consultoria",
                                              "Aporte de Sócio", "Retirada / Pró-labore",
                                              "Outro"], key="em_tipo")
            valor_str = st.text_input("Valor (R$)", value="0,00", placeholder="ex: 5.000,00",
                                       help="Digite o valor no formato 5.000,00", key="em_valor")
            eh_retirada = (tipo_v == "Retirada / Pró-labore")
            valor_v     = parse_brl(valor_str)
            valor_efet  = -valor_v if eh_retirada else valor_v
            rend_v      = round(ultimo_saldo * TAXA_ESCRITORIO, 2)
            saldo_calc  = round(ultimo_saldo + valor_efet + rend_v, 2)
            st.markdown(
                f'<small style="color:#555">'
                f'Rendimento do mês (1% s/ saldo anterior de {brl_md(ultimo_saldo)}): <b>{brl_md(rend_v)}</b>'
                f' &nbsp;·&nbsp; Saldo final: <b>{brl_md(saldo_calc)}</b>'
                f' &nbsp;(= último {brl_md(ultimo_saldo)} '
                f'{"−" if eh_retirada else "+"} valor {brl_md(valor_v)} + rend {brl_md(rend_v)})'
                f'</small>',
                unsafe_allow_html=True)
            if st.button("💾 Salvar Lançamento", key="btn_em", use_container_width=True, type="primary"):
                try:
                    mes_str = mes_v.strftime("%Y-%m-01")
                    sb.table("investimentos_escritorio").upsert({
                        "mes": mes_str, "tipo": tipo_v,
                        "valor": round(valor_efet, 2),
                        "rendimento": rend_v,
                        "saldo_final": saldo_calc,
                        "taxa_mensal": TAXA_ESCRITORIO,
                    }, on_conflict="mes").execute()
                    st.success("Lançamento salvo!")
                    load_investimentos.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro: {e}")

    if inv.empty:
        st.info("Nenhum dado de investimento cadastrado.")
        _form_novo_mes()
        return

    inv_plot = inv[inv["saldo_final"] > 0].copy()
    ultimo   = inv_plot.iloc[-1]
    primeiro = inv_plot.iloc[0]
    saldo_v  = float(ultimo["saldo_final"])
    rend_v   = float(ultimo["rendimento"])
    saldo_ini = float(primeiro["saldo_final"])
    rentab_total = (saldo_v / saldo_ini - 1) * 100 if saldo_ini else 0
    total_aporte = float(inv[inv["tipo"] != "Saldo Inicial"]["valor"].sum())
    total_rend   = float(inv["rendimento"].sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f"""<div class='kpi-card'>
        <div class='kpi-value'>{brl(saldo_v)}</div>
        <div class='kpi-label'>📈 Saldo Atual</div>
    </div>""", unsafe_allow_html=True)
    c2.markdown(f"""<div class='kpi-card ouro'>
        <div class='kpi-value'>{brl(rend_v)}</div>
        <div class='kpi-label'>💰 Rendimento Último Mês</div>
    </div>""", unsafe_allow_html=True)
    c3.markdown(f"""<div class='kpi-card'>
        <div class='kpi-value'>{brl(total_rend)}</div>
        <div class='kpi-label'>💵 Rendimento Acumulado</div>
    </div>""", unsafe_allow_html=True)
    cls_r = "" if rentab_total >= 0 else "neg"
    c4.markdown(f"""<div class='kpi-card {cls_r}'>
        <div class='kpi-value'>{pct(rentab_total)}</div>
        <div class='kpi-label'>📊 Rentabilidade Total</div>
        <div class='kpi-sub'>Desde {primeiro["mes"].strftime("%b/%Y")}</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=inv_plot["mes"].dt.strftime("%b/%Y"), y=inv_plot["rendimento"],
        name="Rendimento", marker_color=OURO, opacity=0.8,
    ))
    fig.add_trace(go.Scatter(
        x=inv_plot["mes"].dt.strftime("%b/%Y"), y=inv_plot["saldo_final"],
        name="Saldo", line=dict(color=VERDE, width=3), mode="lines+markers",
        marker=dict(size=8),
    ))
    fig.update_layout(
        height=300, margin=dict(l=0,r=0,t=10,b=0),
        legend=dict(orientation="h", y=1.12),
        paper_bgcolor="#FFFBEF", plot_bgcolor="#FFFBEF",
        yaxis=dict(tickprefix="R$ ", gridcolor="#EDE8D5"),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)

    df_show = inv_plot[["mes","tipo","valor","rendimento","saldo_final"]].copy()
    df_show.columns = ["Mês","Tipo","Aporte","Rendimento","Saldo Final"]
    df_show["Mês"]        = df_show["Mês"].dt.strftime("%b/%Y")
    df_show["Aporte"]     = df_show["Aporte"].apply(brl)
    df_show["Rendimento"] = df_show["Rendimento"].apply(brl)
    df_show["Saldo Final"]= df_show["Saldo Final"].apply(brl)
    st.dataframe(df_show, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Lançar novo mês ───────────────────────────────────────────────────────
    _form_novo_mes()

    # ── Excluir último lançamento ──────────────────────────────────────────────
    with st.expander("🗑️ Excluir Último Lançamento"):
        ultimo_lanc = inv_plot.iloc[-1]
        st.caption(
            f"Último lançamento: **{ultimo_lanc['mes'].strftime('%b/%Y')}** · "
            f"Tipo: **{ultimo_lanc['tipo']}** · "
            f"Saldo final: **{brl_md(float(ultimo_lanc['saldo_final']))}**"
        )
        if st.session_state.get("confirm_del_escritorio"):
            st.warning(
                f"Excluir o lançamento de {ultimo_lanc['mes'].strftime('%b/%Y')}? "
                f"O saldo do Escritório volta a ser o do mês anterior."
            )
            dc1, dc2 = st.columns(2)
            if dc1.button("✅ Confirmar exclusão", key="conf_del_escritorio",
                          type="primary", use_container_width=True):
                try:
                    mes_str = ultimo_lanc["mes"].strftime("%Y-%m-01")
                    sb.table("investimentos_escritorio").delete().eq("mes", mes_str).execute()
                    st.session_state.pop("confirm_del_escritorio", None)
                    load_investimentos.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao excluir: {e}")
            if dc2.button("❌ Cancelar", key="canc_del_escritorio", use_container_width=True):
                st.session_state.pop("confirm_del_escritorio", None)
                st.rerun()
        else:
            if st.button("🗑️ Excluir Último Lançamento", key="btn_del_escritorio",
                         use_container_width=True):
                st.session_state["confirm_del_escritorio"] = True
                st.rerun()

    # ── Projeção Futura ───────────────────────────────────────────────────────
    with st.expander("📈 Projeção Futura"):
        n_meses_hist = len(inv_plot)
        rend_medio   = (total_rend / n_meses_hist) if n_meses_hist else 0
        taxa_media   = (rend_medio / saldo_v * 100) if saldo_v else 0

        st.caption(f"Rendimento médio histórico: **{brl(rend_medio)}/mês** ({taxa_media:.2f}% a.m.)")
        cp1, cp2, cp3 = st.columns(3)
        meses_proj  = cp1.slider("Meses a projetar", 1, 60, 12, key="proj_meses")
        aporte_proj = cp2.number_input("Aporte mensal (R$)", min_value=0.0,
                                        value=float(total_aporte / n_meses_hist) if n_meses_hist else 0.0,
                                        step=100.0, key="proj_aporte")
        taxa_proj   = cp3.number_input("Taxa mensal (%)", min_value=0.0, max_value=10.0,
                                        value=round(taxa_media, 2), step=0.01,
                                        format="%.2f", key="proj_taxa") / 100

        # Calcular projeção
        saldo_p = saldo_v
        proj_rows = []
        from datetime import date
        mes_base = ultimo["mes"]
        for i in range(1, meses_proj + 1):
            rend_p  = saldo_p * taxa_proj
            saldo_p = saldo_p + aporte_proj + rend_p
            mes_p   = pd.Timestamp(mes_base) + pd.DateOffset(months=i)
            proj_rows.append({"Mês": mes_p.strftime("%b/%Y"),
                               "Aporte": brl(aporte_proj),
                               "Rendimento Est.": brl(rend_p),
                               "Saldo Projetado": brl(saldo_p)})

        if proj_rows:
            df_proj = pd.DataFrame(proj_rows)
            saldo_final_proj = saldo_p
            st.markdown(f"**Saldo projetado em {meses_proj} meses: {brl(saldo_final_proj)}**"
                        f"  ↑ {pct((saldo_final_proj/saldo_v - 1)*100)} vs hoje")

            fig_p = go.Figure(go.Scatter(
                x=df_proj["Mês"], y=[float(r["Saldo Projetado"].replace("R$ ","")
                    .replace(".","").replace(",",".")) for r in proj_rows],
                mode="lines+markers", line=dict(color=VERDE, width=2, dash="dot"),
                marker=dict(size=6), fill="tozeroy", fillcolor="rgba(26,71,49,0.08)",
            ))
            fig_p.update_layout(
                height=220, margin=dict(l=0,r=0,t=10,b=0),
                paper_bgcolor="#FFFBEF", plot_bgcolor="#FFFBEF",
                yaxis=dict(tickprefix="R$ ", gridcolor="#EDE8D5"),
                xaxis=dict(gridcolor="#EDE8D5"),
            )
            st.plotly_chart(fig_p, use_container_width=True)
            st.dataframe(df_proj, use_container_width=True, hide_index=True)


# ─── MAIN ─────────────────────────────────────────────────────────────────────
if "user" not in st.session_state:
    pagina_login()
else:
    render_header()

    # Carga de dados (sem spinner — não interfere com estado de navegação)
    snap_rv       = load_snapshot("RV")
    snap_rf       = load_snapshot("RF")
    posicoes_rv   = load_posicoes_rv()
    historico     = load_historico()
    emprestimos   = load_emprestimos()
    investimentos = load_investimentos()

    # ── Navegação por session_state (nunca perde o estado em reruns) ──────────
    _NAV = [
        "📊 Resumo",
        "📈 Ações BR", "🇧🇷 ETF BR", "🏢 FIIs",
        "🌍 Internacional",
        "🏛️ Tesouro", "📋 CRI/CRA", "💼 Fundos", "🏦 CDB/LCI/LCA",
        "📉 Evolução",
        "💳 Empréstimos", "🏢 Escritório",
    ]
    if "nav_section" not in st.session_state:
        st.session_state["nav_section"] = _NAV[0]

    # CSS: radio horizontal estilizado como abas
    st.markdown("""
    <style>
    div[data-testid="stRadio"] > label { display: none; }
    div[data-testid="stRadio"] div[role="radiogroup"] {
        display: flex; flex-wrap: wrap; gap: 0;
        border-bottom: 2px solid #d0d0d0; margin-bottom: 1.2rem;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label {
        display: flex; align-items: center; padding: 9px 13px;
        cursor: pointer; border-bottom: 3px solid transparent;
        margin-bottom: -2px; color: #555; font-size: 0.82rem;
        white-space: nowrap; background: transparent;
        transition: color .15s;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label:hover {
        color: #1A4731;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label:has(input:checked) {
        border-bottom: 3px solid #1A4731;
        color: #1A4731; font-weight: 700;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] input[type="radio"] {
        display: none;
    }
    </style>""", unsafe_allow_html=True)

    nav = st.radio(
        "Seção", _NAV, horizontal=True,
        key="nav_section", label_visibility="collapsed"
    )

    # ── Conteúdo da seção ativa ───────────────────────────────────────────────
    if   nav == "📊 Resumo":         tab_resumo(snap_rv, snap_rf, posicoes_rv)
    elif nav == "📈 Ações BR":       tab_rv_br("Ações BR",  snap_rv, posicoes_rv)
    elif nav == "🇧🇷 ETF BR":       tab_rv_br("ETF BR",    snap_rv, posicoes_rv)
    elif nav == "🏢 FIIs":           tab_rv_br("FII",       snap_rv, posicoes_rv)
    elif nav == "🌍 Internacional":  tab_internacional(snap_rv, posicoes_rv)
    elif nav == "🏛️ Tesouro":       tab_tesouro(snap_rf)
    elif nav == "📋 CRI/CRA":        tab_cricra(snap_rf)
    elif nav == "💼 Fundos":         tab_fundos(snap_rf)
    elif nav == "🏦 CDB/LCI/LCA":   tab_cdb(snap_rf)
    elif nav == "📉 Evolução":       tab_evolucao(historico)
    elif nav == "💳 Empréstimos":    tab_emprestimos(emprestimos)
    elif nav == "🏢 Escritório":     tab_escritorio(investimentos)

    # Logout
    st.markdown("<br>", unsafe_allow_html=True)
    cols = st.columns([5, 1])
    with cols[1]:
        if st.button("🚪 Sair", use_container_width=True):
            sb.auth.sign_out()
            st.session_state.clear()
            st.rerun()

    st.markdown(
        f"<p style='text-align:center;color:#bbb;font-size:.72rem'>"
        f"3P Finanças · Planejar · Poupar · Prosperar · {agora_br().strftime('%d/%m/%Y %H:%M')}</p>",
        unsafe_allow_html=True,
    )
